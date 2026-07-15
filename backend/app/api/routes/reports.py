from __future__ import annotations

import csv
from decimal import Decimal
import html
import io
import json
import re
import unicodedata
from datetime import date, datetime, timedelta, timezone
from uuid import UUID
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Path, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from xhtml2pdf import pisa
from sqlalchemy import Numeric, Text, case, cast, extract, func, or_, select, update
from sqlalchemy.orm import Session

from app.api.deps import get_db, require_roles
from app.models.catalog import Booking, BookingStatus, CourseSession, CourseType, Location, Professor, SessionStatus
from app.models.client_record import ClientInvoiceLine, ClientManualTransaction, ClientNoteEntry
from app.models.family import ClientFamilyLink
from app.models.ops import CommunicationChannel as CommunicationChannelModel, CommunicationLog, LegalEntity
from app.models.payout import ProfessorSessionPayout
from app.models.product_catalog import CatalogKit, CatalogKitItem, CatalogProduct, ProductCategory, ProductLocationStock
from app.models.quote import Prospect, Quote, QuoteLine
from app.models.reporting import GeneratedReport
from app.models.typeform_intake import TypeformFormConfig, TypeformIntake
from app.models.user import User, UserRole
from app.schemas.report import (
    AttendanceReportRow,
    CommunicationChannel,
    CommunicationFiltersOut,
    CommunicationPeriod,
    CommunicationReportPageOut,
    CommunicationProfessorFilterOut,
    CommunicationReportRow,
    CommunicationResendRequest,
    CommunicationTypeFilterOut,
    GeneratedReportCreate,
    GeneratedReportOut,
    IntakeFamilyChildSummary,
    IntakeFamilySummaryRow,
    ProfessorStatementRow,
    ReservationReportRow,
)
from app.services.communication_journal import COMMUNICATION_TYPE_LABELS, KNOWN_COMMUNICATION_TYPES, communication_type_label
from app.services.email_delivery import email_delivery_disabled_reason, send_email
from app.services.messaging_templates import resolve_sender_profile

router = APIRouter(prefix="/admin/reports")
INVOICE_RANGE_NOTE_PREFIX = "INVOICE_RANGE::"
COMMUNICATION_ARCHIVE_RETENTION_DAYS = 365
ADMIN_COMMUNICATION_TIMEZONE = ZoneInfo("Europe/Paris")
REPORT_TYPE_LABELS: dict[str, str] = {
    "intake-families": "Synthese intakes par famille",
    "quote-families": "Synthese devis par famille",
    "expired-quotes": "Devis expires/refuses/annules",
    "overdue-invoices": "Factures echues non payees",
    "reservations": "Reservations",
    "attendance": "Presence eleves",
    "professor-statements": "Releves professeurs",
    "communications": "Communications",
    "payments": "Paiements clients",
    "quotes": "Devis",
    "subscriptions": "Abonnements",
    "planning-fill": "Remplissage planning",
    "check-deposits": "Depots de cheques",
    "material-forecast": "Approvisionnement partitions et jeux de notes",
    "referrals": "Parrainages",
    "teacher-payments": "Paiement des salaires",
}


def _professor_name(prof: Professor) -> str:
    return f"{prof.first_name} {prof.last_name}".strip()


def _client_name(user: User) -> str:
    value = f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip()
    return value or user.email


def _service_address(user: User) -> str:
    parts = [
        (user.address_line or "").strip(),
        (user.postal_code or "").strip(),
        (user.city or "").strip(),
        (user.address_country or "").strip().upper(),
    ]
    return " ".join(part for part in parts if part).strip()


def _invoice_fields_from_note_message(message: str | None) -> tuple[str | None, str | None]:
    raw = (message or "").strip()
    if not raw:
        return None, None
    prefix_index = raw.find(INVOICE_RANGE_NOTE_PREFIX)
    if prefix_index >= 0:
        payload = raw[prefix_index + len(INVOICE_RANGE_NOTE_PREFIX) :].strip()
        if payload:
            try:
                parsed = json.loads(payload)
                if isinstance(parsed, dict):
                    invoice_number = str(parsed.get("invoice_number") or "").strip() or None
                    invoice_status = str(parsed.get("invoice_status") or "").strip().upper() or None
                    return invoice_number, invoice_status
            except json.JSONDecodeError:
                pass
    # Fallback for legacy note text if JSON payload is not present.
    match = re.search(r"\bFacture\s+([A-Za-z0-9._/-]+)", raw)
    invoice_number = match.group(1).strip() if match is not None else None
    return invoice_number, None


def _invoice_range_metadata_from_note_message(message: str | None) -> dict[str, object] | None:
    raw = (message or "").strip()
    if not raw:
        return None
    prefix_index = raw.find(INVOICE_RANGE_NOTE_PREFIX)
    if prefix_index < 0:
        return None
    payload = raw[prefix_index + len(INVOICE_RANGE_NOTE_PREFIX) :].strip()
    if not payload:
        return None
    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError:
        return None
    if not isinstance(parsed, dict):
        return None
    if str(parsed.get("kind") or "").strip().upper() != "INVOICE_RANGE":
        return None
    return parsed


def _ensure_date_range(from_: datetime | None, to: datetime | None) -> None:
    if from_ is not None and to is not None and from_ > to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'from' must be before 'to'",
        )


def _day_bounds(day: date, tz: ZoneInfo = ADMIN_COMMUNICATION_TIMEZONE) -> tuple[datetime, datetime]:
    start_local = datetime(day.year, day.month, day.day, tzinfo=tz)
    end_local = start_local + timedelta(days=1)
    return start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc)


def _communication_period_bounds(
    period: CommunicationPeriod,
    now_utc: datetime,
) -> tuple[datetime, datetime] | None:
    today_start, today_end = _day_bounds(now_utc.astimezone(ADMIN_COMMUNICATION_TIMEZONE).date())
    if period == CommunicationPeriod.ALL:
        return None
    if period == CommunicationPeriod.TODAY:
        return today_start, today_end
    if period == CommunicationPeriod.WEEK:
        return now_utc - timedelta(days=7), now_utc
    if period == CommunicationPeriod.MONTH:
        return now_utc - timedelta(days=30), now_utc
    if period == CommunicationPeriod.SEMESTER:
        return now_utc - timedelta(days=183), now_utc
    if period == CommunicationPeriod.YEAR:
        return now_utc - timedelta(days=365), now_utc
    return today_start, today_end


def _archive_communications_older_than_one_year(db: Session, now_utc: datetime) -> None:
    archive_before = now_utc - timedelta(days=COMMUNICATION_ARCHIVE_RETENTION_DAYS)
    db.execute(
        update(CommunicationLog)
        .where(
            CommunicationLog.archived_at.is_(None),
            CommunicationLog.occurred_at < archive_before,
        )
        .values(archived_at=now_utc, updated_at=now_utc)
    )
    db.commit()


def _parse_communication_log_id(raw_value: str) -> UUID:
    raw = str(raw_value or "").strip()
    if raw.startswith("communication-log-"):
        raw = raw[len("communication-log-") :]
    try:
        return UUID(raw)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Identifiant de communication invalide",
        ) from exc


def _communication_row_out(row: CommunicationLog) -> CommunicationReportRow:
    return CommunicationReportRow(
        id=f"communication-log-{row.id}",
        channel=CommunicationChannel(row.channel.value),
        source=row.source,
        communication_type=row.communication_type,
        communication_type_label=communication_type_label(row.communication_type),
        sender_category=row.sender_category.value,
        sender_label=row.sender_label,
        sender_user_id=row.sender_user_id,
        professor_id=row.professor_id,
        occurred_at=row.occurred_at,
        subject=row.subject,
        recipient=row.recipient,
        recipient_user_id=row.recipient_user_id,
        delivery_status=row.delivery_status.value,
        provider_message_id=row.provider_message_id,
        provider=row.provider,
        content=row.content,
        content_format=row.content_format.value,
        error_message=row.error_message,
    )


def _text(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _json_object(value: object | None) -> dict[str, object]:
    return value if isinstance(value, dict) else {}


def _json_list(value: object | None) -> list[object]:
    return value if isinstance(value, list) else []


def _bool_or_default(value: object | None, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    lowered = _text(value).casefold()
    if lowered in {"1", "true", "yes", "on", "oui"}:
        return True
    if lowered in {"0", "false", "no", "off", "non"}:
        return False
    return default


def _normalize_token(value: object | None) -> str:
    raw = unicodedata.normalize("NFKD", _text(value))
    ascii_text = raw.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", ascii_text.casefold())


def _display_name(first_name: object | None, last_name: object | None, fallback: str = "-") -> str:
    label = " ".join(part for part in [_text(first_name), _text(last_name)] if part).strip()
    return label or fallback


MONTH_LABELS_FR = [
    "janvier",
    "fevrier",
    "mars",
    "avril",
    "mai",
    "juin",
    "juillet",
    "aout",
    "septembre",
    "octobre",
    "novembre",
    "decembre",
]
MONTH_NUMBER_BY_TOKEN = {label: index + 1 for index, label in enumerate(MONTH_LABELS_FR)}


def _month_year_label(month: int, year: int) -> str:
    return f"{MONTH_LABELS_FR[month - 1]} {year}"


def _manual_payment_method_code(reference: str | None) -> str | None:
    normalized_reference = (reference or "").strip()
    if not normalized_reference.upper().startswith("MODE:"):
        return None
    suffix = normalized_reference[5:].strip()
    if not suffix:
        return None
    separator_index = suffix.upper().find("|REF:")
    return (suffix[:separator_index] if separator_index >= 0 else suffix).strip().upper() or None


def _deposit_month_year_from_text(value: str | None) -> tuple[int, int] | None:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii").casefold()
    ascii_text = re.sub(r"[^a-z0-9]+", " ", ascii_text)
    pattern = r"\b(" + "|".join(MONTH_NUMBER_BY_TOKEN.keys()) + r")\s+((?:20|19)\d{2})\b"
    matches = list(re.finditer(pattern, ascii_text))
    if not matches:
        return None
    for match in matches:
        prefix = ascii_text[max(0, match.start() - 40) : match.start()]
        if "depos" in prefix or "banque" in prefix:
            return MONTH_NUMBER_BY_TOKEN[match.group(1)], int(match.group(2))
    match = matches[-1]
    return MONTH_NUMBER_BY_TOKEN[match.group(1)], int(match.group(2))


def _is_legacy_demian_check(transaction: ClientManualTransaction, client: User) -> bool:
    amount = abs(Decimal(transaction.total_incl_vat or 0)).quantize(Decimal("0.01"))
    return (
        _normalize_token(client.first_name) == "myriam"
        and _normalize_token(client.last_name) == "demian"
        and (client.email or "").strip().casefold() == "myriamthera@hotmail.com"
        and amount == Decimal("616.00")
        and _manual_payment_method_code(transaction.reference) == "CHECK"
    )


def _legacy_demian_check_deposit_months(
    rows: list[tuple[ClientManualTransaction, User]],
) -> dict[UUID, tuple[int, int]]:
    target_months = [(9, 2026), (2, 2027)]
    candidates = [
        (transaction, client)
        for transaction, client in rows
        if _is_legacy_demian_check(transaction, client)
    ]
    used_months: set[tuple[int, int]] = set()
    for transaction, _client in candidates:
        parsed = (
            _deposit_month_year_from_text(transaction.description)
            or _deposit_month_year_from_text(transaction.label)
        )
        if parsed in target_months:
            used_months.add(parsed)
    remaining_months = [target for target in target_months if target not in used_months]
    unlabeled_candidates = [
        transaction
        for transaction, _client in candidates
        if not (
            _deposit_month_year_from_text(transaction.description)
            or _deposit_month_year_from_text(transaction.label)
        )
    ]
    unlabeled_candidates.sort(key=lambda transaction: (transaction.created_at, str(transaction.id)))
    return {
        transaction.id: month_year
        for transaction, month_year in zip(unlabeled_candidates, remaining_months, strict=False)
    }


def _local_date_label(value: datetime | None) -> str:
    if value is None:
        return "-"
    return value.astimezone(ADMIN_COMMUNICATION_TIMEZONE).strftime("%d/%m/%Y")


def _amount_label(value: Decimal) -> str:
    amount = value.quantize(Decimal("0.01"))
    return f"{amount:,.2f}".replace(",", " ").replace(".", ",")


def _check_deposit_students_for_transaction(
    db: Session,
    *,
    client_id: UUID,
    student_user_id: UUID | None,
) -> list[User]:
    if student_user_id is not None:
        student = db.get(User, student_user_id)
        return [student] if student is not None else []
    return list(
        db.scalars(
            select(User)
            .join(ClientFamilyLink, ClientFamilyLink.child_user_id == User.id)
            .where(ClientFamilyLink.adult_user_id == client_id)
            .order_by(User.last_name.asc().nulls_last(), User.first_name.asc().nulls_last())
        ).all()
    )


def _check_deposit_report_rows(
    db: Session,
    *,
    month: int,
    year: int,
    legal_entity_id: UUID,
) -> tuple[LegalEntity, list[dict[str, object]]]:
    legal_entity = db.get(LegalEntity, legal_entity_id)
    if legal_entity is None or not legal_entity.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entite legale introuvable")

    rows = db.execute(
        select(ClientManualTransaction, User)
        .join(User, User.id == ClientManualTransaction.user_id)
        .where(
            ClientManualTransaction.transaction_type == "PAYMENT",
            ClientManualTransaction.status == "CHECK_RECEIVED",
            ClientManualTransaction.reference.ilike("%MODE:CHECK%"),
            or_(
                ClientManualTransaction.legal_entity_id == legal_entity_id,
                func.lower(User.email) == "myriamthera@hotmail.com",
            ),
        )
        .order_by(ClientManualTransaction.occurred_at.asc(), User.last_name.asc().nulls_last(), User.first_name.asc().nulls_last())
    ).all()
    selected_is_piano_academie = _normalize_token(legal_entity.name) == "pianoacademie"
    legacy_demian_deposit_months = _legacy_demian_check_deposit_months(rows) if selected_is_piano_academie else {}
    report_rows: list[dict[str, object]] = []
    for transaction, client in rows:
        legal_entity_matches = transaction.legal_entity_id == legal_entity_id
        legacy_entity_match = selected_is_piano_academie and transaction.id in legacy_demian_deposit_months
        if not legal_entity_matches and not legacy_entity_match:
            continue
        deposit_month_year = (
            _deposit_month_year_from_text(transaction.description)
            or _deposit_month_year_from_text(transaction.label)
            or legacy_demian_deposit_months.get(transaction.id)
        )
        if deposit_month_year != (month, year):
            continue
        students = _check_deposit_students_for_transaction(
            db,
            client_id=client.id,
            student_user_id=transaction.student_user_id,
        )
        student_names = [_display_name(student.first_name, student.last_name, "") for student in students]
        report_rows.append(
            {
                "responsible_last_name": _text(client.last_name),
                "responsible_first_name": _text(client.first_name),
                "student_last_name": ", ".join(_text(student.last_name) for student in students if _text(student.last_name)),
                "student_first_name": ", ".join(_text(student.first_name) for student in students if _text(student.first_name)),
                "student_display": ", ".join(name for name in student_names if name),
                "received_at": _local_date_label(transaction.occurred_at),
                "deposit_label": _month_year_label(month, year),
                "amount": abs(Decimal(transaction.total_incl_vat or 0)).quantize(Decimal("0.01")),
            }
        )
    return legal_entity, report_rows


def _check_deposit_report_row_json(row: dict[str, object]) -> dict[str, object]:
    amount = row.get("amount")
    return {
        "responsible_last_name": _text(row.get("responsible_last_name")),
        "responsible_first_name": _text(row.get("responsible_first_name")),
        "student_last_name": _text(row.get("student_last_name")),
        "student_first_name": _text(row.get("student_first_name")),
        "received_at": _text(row.get("received_at")),
        "deposit_label": _text(row.get("deposit_label")),
        "amount": str(amount if isinstance(amount, Decimal) else Decimal("0.00")),
    }


def _check_deposit_report_rows_from_content(row: GeneratedReport) -> tuple[int, int, str, list[dict[str, object]]]:
    content = _json_object(row.content_json)
    month = int(content.get("month") or 1)
    year = int(content.get("year") or datetime.now(ADMIN_COMMUNICATION_TIMEZONE).year)
    legal_entity_name = _text(content.get("legal_entity_name")) or "-"
    items = [_json_object(item) for item in _json_list(content.get("items"))]
    rows: list[dict[str, object]] = []
    for item in items:
        rows.append(
            {
                "responsible_last_name": _text(item.get("responsible_last_name")),
                "responsible_first_name": _text(item.get("responsible_first_name")),
                "student_last_name": _text(item.get("student_last_name")),
                "student_first_name": _text(item.get("student_first_name")),
                "received_at": _text(item.get("received_at")),
                "deposit_label": _text(item.get("deposit_label")),
                "amount": Decimal(_text(item.get("amount")) or "0.00"),
            }
        )
    return month, year, legal_entity_name, rows


def _render_check_deposit_report_pdf(
    *,
    rows: list[dict[str, object]],
    month: int,
    year: int,
    legal_entity: LegalEntity,
) -> bytes:
    title = f"Cheques a deposer - {_month_year_label(month, year)} - {legal_entity.name}"
    body_rows = "\n".join(
        "<tr>"
        f"<td>{html.escape(_text(row.get('responsible_last_name')))}</td>"
        f"<td>{html.escape(_text(row.get('responsible_first_name')))}</td>"
        f"<td>{html.escape(_text(row.get('student_last_name')))}</td>"
        f"<td>{html.escape(_text(row.get('student_first_name')))}</td>"
        f"<td>{html.escape(_text(row.get('received_at')))}</td>"
        f"<td>{html.escape(_text(row.get('deposit_label')))}</td>"
        f"<td class='amount'>{html.escape(_amount_label(row.get('amount') if isinstance(row.get('amount'), Decimal) else Decimal('0')))} EUR</td>"
        "<td>&nbsp;</td>"
        "</tr>"
        for row in rows
    )
    if not body_rows:
        body_rows = "<tr><td colspan='8' class='empty'>Aucun cheque recu a deposer pour cette periode.</td></tr>"
    document = f"""
    <html>
      <head>
        <meta charset="utf-8" />
        <style>
          @page {{ size: A4 landscape; margin: 18mm; }}
          body {{ font-family: Helvetica, Arial, sans-serif; color: #1f2937; font-size: 10px; }}
          h1 {{ font-size: 18px; margin: 0 0 6px; }}
          p {{ margin: 0 0 16px; color: #6b7280; }}
          table {{ width: 100%; border-collapse: collapse; }}
          th {{ background: #eef2f7; font-weight: bold; }}
          th, td {{ border: 1px solid #cbd5e1; padding: 7px 6px; vertical-align: top; }}
          td.amount {{ text-align: right; white-space: nowrap; }}
          td.empty {{ text-align: center; color: #6b7280; padding: 18px; }}
        </style>
      </head>
      <body>
        <h1>{html.escape(title)}</h1>
        <p>Liste des cheques au statut recu, filtres par mois de depot prevu et entite legale.</p>
        <table>
          <thead>
            <tr>
              <th>Nom responsable</th>
              <th>Prenom responsable</th>
              <th>Nom eleve</th>
              <th>Prenom eleve</th>
              <th>Date reception cheque</th>
              <th>Date prevue depot en banque</th>
              <th>Montant cheque</th>
              <th>Date depot banque</th>
            </tr>
          </thead>
          <tbody>{body_rows}</tbody>
        </table>
      </body>
    </html>
    """
    output = io.BytesIO()
    result = pisa.CreatePDF(io.StringIO(document), dest=output, encoding="utf-8")
    if result.err:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Generation PDF impossible")
    return output.getvalue()


def _render_stored_check_deposit_report_pdf(row: GeneratedReport) -> bytes:
    month, year, legal_entity_name, rows = _check_deposit_report_rows_from_content(row)

    class StoredLegalEntity:
        name = legal_entity_name

    return _render_check_deposit_report_pdf(
        rows=rows,
        month=month,
        year=year,
        legal_entity=StoredLegalEntity(),
    )


def _render_check_deposit_report_xlsx(
    *,
    rows: list[dict[str, object]],
    month: int,
    year: int,
    legal_entity: LegalEntity,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cheques a deposer"
    worksheet.append([f"Cheques a deposer - {_month_year_label(month, year)} - {legal_entity.name}"])
    worksheet.append([])
    headers = [
        "Nom responsable",
        "Prenom responsable",
        "Nom eleve",
        "Prenom eleve",
        "Date reception cheque",
        "Date prevue depot en banque",
        "Montant cheque",
        "Date depot banque",
    ]
    worksheet.append(headers)
    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2F7")
    for row in rows:
        worksheet.append(
            [
                _text(row.get("responsible_last_name")),
                _text(row.get("responsible_first_name")),
                _text(row.get("student_last_name")),
                _text(row.get("student_first_name")),
                _text(row.get("received_at")),
                _text(row.get("deposit_label")),
                float(row.get("amount") if isinstance(row.get("amount"), Decimal) else Decimal("0")),
                "",
            ]
        )
    worksheet.freeze_panes = "A4"
    widths = [22, 22, 24, 24, 22, 28, 16, 22]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for cell in worksheet["G"][3:]:
        cell.number_format = '#,##0.00 "EUR"'
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _render_stored_check_deposit_report_xlsx(row: GeneratedReport) -> bytes:
    month, year, legal_entity_name, rows = _check_deposit_report_rows_from_content(row)

    class StoredLegalEntity:
        name = legal_entity_name

    return _render_check_deposit_report_xlsx(
        rows=rows,
        month=month,
        year=year,
        legal_entity=StoredLegalEntity(),
    )


PARIS_STOCK_LOCATION_CODES = {"ASSAS", "DOMICILE", "DULONG", "POMPE", "RICHELIEU", "SCHEFFER"}


def _material_report_school_year_bounds(school_year_label: str | None) -> tuple[date | None, date | None]:
    raw = _text(school_year_label)
    match = re.fullmatch(r"(\d{4})-(\d{4})", raw)
    if match is None:
        return None, None
    start_year = int(match.group(1))
    end_year = int(match.group(2))
    return date(start_year, 9, 1), date(end_year, 8, 31)


def _quote_calendar_location_ids(quote: Quote) -> list[UUID]:
    location_ids: list[UUID] = []
    if quote.location_id is not None:
        location_ids.append(quote.location_id)
    calendar = _json_object(quote.calendar_snapshot)
    for section_name in ("blocks", "sessions"):
        for raw_item in _json_list(calendar.get(section_name)):
            item = _json_object(raw_item)
            raw_id = _text(item.get("location_id"))
            if not raw_id:
                continue
            try:
                location_id = UUID(raw_id)
            except ValueError:
                continue
            if location_id not in location_ids:
                location_ids.append(location_id)
    solfege_slot = _json_object(_json_object(calendar.get("solfege")).get("selected_slot"))
    raw_solfege_location_id = _text(solfege_slot.get("location_id"))
    if raw_solfege_location_id:
        try:
            solfege_location_id = UUID(raw_solfege_location_id)
            if solfege_location_id not in location_ids:
                location_ids.append(solfege_location_id)
        except ValueError:
            pass
    return location_ids


def _material_report_site_for_location(location: Location | None) -> str | None:
    if location is None:
        return None
    code = _text(location.code).upper()
    city = _normalize_token(location.city)
    label = _normalize_token(f"{location.name} {location.code} {location.city}")
    if code == "BAR_LE_DUC" or "barleduc" in label:
        return "BAR_LE_DUC"
    if code in PARIS_STOCK_LOCATION_CODES or city == "paris":
        return "PARIS"
    if bool(location.is_online):
        return "PARIS"
    return None


def _material_report_stock_location_site(location: Location | None) -> str | None:
    if location is None:
        return None
    code = _text(location.code).upper()
    if code == "BAR_LE_DUC":
        return "BAR_LE_DUC"
    if code in PARIS_STOCK_LOCATION_CODES:
        return "PARIS"
    return None


def _material_report_site_for_quote(quote: Quote, locations_by_id: dict[UUID, Location]) -> str | None:
    fallback: str | None = None
    for location_id in _quote_calendar_location_ids(quote):
        location = locations_by_id.get(location_id)
        site = _material_report_site_for_location(location)
        if site is None:
            continue
        if location is not None and not bool(location.is_online):
            return site
        fallback = fallback or site
    return fallback


def _material_report_site_label(site_key: str) -> str:
    if site_key == "BAR_LE_DUC":
        return "Bar-le-Duc"
    if site_key == "PARIS":
        return "Paris"
    return "Non classe"


def _material_product_kind(category_name: object | None, product_title: object | None) -> str | None:
    category_token = _normalize_token(category_name)
    title_token = _normalize_token(product_title)
    if "partition" in category_token or "partition" in title_token:
        return "Partition"
    if "jeudenotes" in title_token:
        return "Jeu de notes"
    return None


def _material_quantity(value: object | None) -> float:
    try:
        quantity = Decimal(str(value if value is not None else "0"))
    except Exception:
        return 0.0
    return float(quantity)


def _material_report_product_json(
    product: CatalogProduct | None,
    category_name: str | None,
) -> dict[str, object]:
    return {
        "product_id": str(product.id) if product is not None else None,
        "product_title": product.title if product is not None else "Produit supprime",
        "category_name": category_name or "-",
    }


def _build_material_forecast_report(
    db: Session,
    *,
    school_year_label: str | None,
    approved_from: date | None = None,
    approved_to: date | None = None,
    status_filter: str | None = None,
    q: str | None = None,
) -> dict[str, object]:
    if approved_from is not None and approved_to is not None and approved_from > approved_to:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="'approved_from' must be before 'approved_to'")

    normalized_status = _text(status_filter).casefold() or "approved"
    stmt = (
        select(Quote, Prospect, User)
        .outerjoin(Prospect, Prospect.id == Quote.prospect_id)
        .outerjoin(User, User.id == Quote.client_id)
        .where(Quote.status == normalized_status)
        .order_by(Quote.approved_at.desc().nullslast(), Quote.created_at.desc())
        .limit(10000)
    )
    if school_year_label:
        stmt = stmt.where(Quote.school_year_label == school_year_label)
    if approved_from is not None:
        start_local, _ = _day_bounds(approved_from)
        stmt = stmt.where(func.coalesce(Quote.approved_at, Quote.updated_at) >= start_local)
    if approved_to is not None:
        _, end_local = _day_bounds(approved_to)
        stmt = stmt.where(func.coalesce(Quote.approved_at, Quote.updated_at) < end_local)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Quote.quote_number.ilike(like),
                Quote.school_year_label.ilike(like),
                cast(Quote.meta, Text).ilike(like),
                cast(Quote.calendar_snapshot, Text).ilike(like),
                Prospect.first_name.ilike(like),
                Prospect.last_name.ilike(like),
                Prospect.email.ilike(like),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
                User.email.ilike(like),
            )
        )

    quote_rows = db.execute(stmt).all()
    quote_ids = [quote.id for quote, _, _ in quote_rows]
    all_location_ids: set[UUID] = set()
    for quote, _, _ in quote_rows:
        all_location_ids.update(_quote_calendar_location_ids(quote))
    locations_by_id: dict[UUID, Location] = {}
    if all_location_ids:
        locations_by_id = {
            location.id: location
            for location in db.scalars(select(Location).where(Location.id.in_(list(all_location_ids)))).all()
        }
    all_locations = db.scalars(select(Location)).all()
    stock_location_ids_by_site: dict[str, set[UUID]] = {"PARIS": set(), "BAR_LE_DUC": set()}
    for location in all_locations:
        site = _material_report_stock_location_site(location)
        if site in stock_location_ids_by_site:
            stock_location_ids_by_site[site].add(location.id)

    lines_by_quote_id: dict[UUID, list[QuoteLine]] = {quote_id: [] for quote_id in quote_ids}
    direct_product_ids: set[UUID] = set()
    kit_ids: set[UUID] = set()
    if quote_ids:
        quote_lines = db.scalars(
            select(QuoteLine)
            .where(QuoteLine.quote_id.in_(quote_ids))
            .order_by(QuoteLine.quote_id.asc(), QuoteLine.sort_order.asc(), QuoteLine.created_at.asc())
        ).all()
        for line in quote_lines:
            lines_by_quote_id.setdefault(line.quote_id, []).append(line)
            if line.product_id is not None:
                direct_product_ids.add(line.product_id)
            if line.kit_id is not None:
                kit_ids.add(line.kit_id)

    kit_items_by_kit_id: dict[UUID, list[CatalogKitItem]] = {kit_id: [] for kit_id in kit_ids}
    kits_by_id: dict[UUID, CatalogKit] = {}
    kit_product_ids: set[UUID] = set()
    if kit_ids:
        kits_by_id = {kit.id: kit for kit in db.scalars(select(CatalogKit).where(CatalogKit.id.in_(list(kit_ids)))).all()}
        for item in db.scalars(
            select(CatalogKitItem)
            .where(CatalogKitItem.kit_id.in_(list(kit_ids)))
            .order_by(CatalogKitItem.kit_id.asc(), CatalogKitItem.display_order.asc(), CatalogKitItem.created_at.asc())
        ).all():
            kit_items_by_kit_id.setdefault(item.kit_id, []).append(item)
            kit_product_ids.add(item.product_id)

    product_ids = direct_product_ids | kit_product_ids
    products_by_id: dict[UUID, CatalogProduct] = {}
    category_by_product_id: dict[UUID, str | None] = {}
    if product_ids:
        product_rows = db.execute(
            select(CatalogProduct, ProductCategory.name)
            .outerjoin(ProductCategory, ProductCategory.id == CatalogProduct.category_id)
            .where(CatalogProduct.id.in_(list(product_ids)))
        ).all()
        for product, category_name in product_rows:
            products_by_id[product.id] = product
            category_by_product_id[product.id] = category_name

    stock_by_site_product: dict[tuple[str, UUID], int] = {}
    if product_ids:
        stock_rows = db.execute(
            select(
                ProductLocationStock.product_id,
                ProductLocationStock.location_id,
                ProductLocationStock.real_quantity,
            ).where(ProductLocationStock.product_id.in_(list(product_ids)))
        ).all()
        for product_id, location_id, quantity in stock_rows:
            for site, site_location_ids in stock_location_ids_by_site.items():
                if location_id in site_location_ids:
                    key = (site, product_id)
                    stock_by_site_product[key] = stock_by_site_product.get(key, 0) + int(quantity or 0)

    summary_by_key: dict[tuple[str, str], dict[str, object]] = {}
    details: list[dict[str, object]] = []

    def add_quantity(
        *,
        site: str,
        quote: Quote,
        prospect: Prospect | None,
        client: User | None,
        source: str,
        product: CatalogProduct | None,
        category_name: str | None,
        product_title: str,
        kind: str,
        quantity: float,
        kit: CatalogKit | None = None,
        quote_line_title: str | None = None,
    ) -> None:
        if quantity <= 0:
            return
        product_key = str(product.id) if product is not None else f"missing:{product_title}"
        key = (site, product_key)
        stock_quantity = stock_by_site_product.get((site, product.id), 0) if product is not None else 0
        if key not in summary_by_key:
            summary_by_key[key] = {
                **_material_report_product_json(product, category_name),
                "site": site,
                "site_label": _material_report_site_label(site),
                "kind": kind,
                "expected_direct": 0.0,
                "expected_from_kits": 0.0,
                "expected_total": 0.0,
                "stock_quantity": stock_quantity,
                "to_order": 0.0,
            }
        summary = summary_by_key[key]
        if source == "kit":
            summary["expected_from_kits"] = float(summary.get("expected_from_kits") or 0) + quantity
        else:
            summary["expected_direct"] = float(summary.get("expected_direct") or 0) + quantity
        expected_total = float(summary.get("expected_direct") or 0) + float(summary.get("expected_from_kits") or 0)
        summary["expected_total"] = expected_total
        summary["to_order"] = max(expected_total - float(summary.get("stock_quantity") or 0), 0)
        details.append(
            {
                "site": site,
                "site_label": _material_report_site_label(site),
                "quote_id": str(quote.id),
                "quote_number": quote.quote_number,
                "approved_at": quote.approved_at.isoformat() if quote.approved_at else None,
                "student_name": _quote_student_name(quote, prospect, client),
                "source": "Kit inscription" if source == "kit" else "Ligne devis",
                "kit_title": kit.title if kit is not None else "",
                "quote_line_title": quote_line_title or "",
                "kind": kind,
                "category_name": category_name or "-",
                "product_id": str(product.id) if product is not None else None,
                "product_title": product_title,
                "quantity": quantity,
            }
        )

    for quote, prospect, client in quote_rows:
        site = _material_report_site_for_quote(quote, locations_by_id)
        if site not in {"PARIS", "BAR_LE_DUC"}:
            continue
        for line in lines_by_quote_id.get(quote.id, []):
            line_quantity = _material_quantity(line.quantity)
            if line.product_id is not None and line.kit_id is None:
                product = products_by_id.get(line.product_id)
                category_name = category_by_product_id.get(line.product_id)
                product_title = product.title if product is not None else line.title
                kind = _material_product_kind(category_name, product_title)
                if kind is not None:
                    add_quantity(
                        site=site,
                        quote=quote,
                        prospect=prospect,
                        client=client,
                        source="direct",
                        product=product,
                        category_name=category_name,
                        product_title=product_title,
                        kind=kind,
                        quantity=line_quantity,
                        quote_line_title=line.title,
                    )
            if line.kit_id is None:
                continue
            kit = kits_by_id.get(line.kit_id)
            for item in kit_items_by_kit_id.get(line.kit_id, []):
                product = products_by_id.get(item.product_id)
                category_name = category_by_product_id.get(item.product_id)
                product_title = product.title if product is not None else "Produit supprime"
                kind = _material_product_kind(category_name, product_title)
                if kind is None:
                    continue
                add_quantity(
                    site=site,
                    quote=quote,
                    prospect=prospect,
                    client=client,
                    source="kit",
                    product=product,
                    category_name=category_name,
                    product_title=product_title,
                    kind=kind,
                    quantity=line_quantity * float(item.quantity or 0),
                    kit=kit,
                    quote_line_title=line.title,
                )

    summary_rows = sorted(
        summary_by_key.values(),
        key=lambda row: (
            _text(row.get("site_label")),
            _text(row.get("kind")),
            _text(row.get("category_name")),
            _text(row.get("product_title")).casefold(),
        ),
    )
    details.sort(
        key=lambda row: (
            _text(row.get("site_label")),
            _text(row.get("product_title")).casefold(),
            _text(row.get("quote_number")),
        )
    )
    return {
        "school_year_label": school_year_label,
        "status": normalized_status,
        "summary_rows": summary_rows,
        "details": details,
        "quote_count": len(quote_rows),
    }


def _quantity_cell_value(value: object | None) -> int | float:
    quantity = _material_quantity(value)
    if quantity.is_integer():
        return int(quantity)
    return quantity


def _append_material_summary_sheet(workbook: Workbook, *, title: str, site: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append([f"Approvisionnement partitions et jeux de notes - {_material_report_site_label(site)}"])
    worksheet.append(["Le stock peut etre complete dans la colonne H ; la colonne I calcule le reste a commander."])
    worksheet.append([])
    headers = [
        "Site",
        "Nature",
        "Categorie catalogue",
        "Nom partition / produit",
        "Nombre attendu",
        "Dont lignes devis",
        "Dont kits inscription",
        "Nombre en stock",
        "Nombre a commander",
        "Produit ID",
    ]
    worksheet.append(headers)
    for cell in worksheet[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2F7")
    filtered_rows = [row for row in rows if _text(row.get("site")) == site]
    for row in filtered_rows:
        excel_row = worksheet.max_row + 1
        worksheet.append(
            [
                _text(row.get("site_label")),
                _text(row.get("kind")),
                _text(row.get("category_name")),
                _text(row.get("product_title")),
                _quantity_cell_value(row.get("expected_total")),
                _quantity_cell_value(row.get("expected_direct")),
                _quantity_cell_value(row.get("expected_from_kits")),
                _quantity_cell_value(row.get("stock_quantity")),
                f"=MAX(E{excel_row}-H{excel_row},0)",
                _text(row.get("product_id")),
            ]
        )
    if not filtered_rows:
        worksheet.append([_material_report_site_label(site), "-", "-", "Aucune partition ou jeu de notes attendu", 0, 0, 0, 0, 0, ""])
    worksheet.freeze_panes = "A5"
    widths = [16, 16, 22, 42, 18, 18, 20, 18, 20, 38]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def _append_material_detail_sheet(workbook: Workbook, *, title: str, site: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append([f"Detail par devis - {_material_report_site_label(site)}"])
    worksheet.append([])
    headers = [
        "Site",
        "Devis",
        "Date validation",
        "Eleve",
        "Origine",
        "Kit",
        "Ligne devis",
        "Nature",
        "Categorie catalogue",
        "Nom partition / produit",
        "Quantite",
        "Produit ID",
    ]
    worksheet.append(headers)
    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2F7")
    filtered_rows = [row for row in rows if _text(row.get("site")) == site]
    for row in filtered_rows:
        worksheet.append(
            [
                _text(row.get("site_label")),
                _text(row.get("quote_number")),
                _text(row.get("approved_at"))[:10],
                _text(row.get("student_name")),
                _text(row.get("source")),
                _text(row.get("kit_title")),
                _text(row.get("quote_line_title")),
                _text(row.get("kind")),
                _text(row.get("category_name")),
                _text(row.get("product_title")),
                _quantity_cell_value(row.get("quantity")),
                _text(row.get("product_id")),
            ]
        )
    if not filtered_rows:
        worksheet.append([_material_report_site_label(site), "-", "-", "-", "-", "-", "-", "-", "-", "Aucune ligne", 0, ""])
    worksheet.freeze_panes = "A4"
    widths = [16, 24, 16, 28, 18, 34, 34, 16, 22, 42, 12, 38]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def _render_material_forecast_report_xlsx(row: GeneratedReport) -> bytes:
    content = _json_object(row.content_json)
    summary_rows = [_json_object(item) for item in _json_list(content.get("summary_rows"))]
    detail_rows = [_json_object(item) for item in _json_list(content.get("details"))]
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    _append_material_summary_sheet(workbook, title="Paris", site="PARIS", rows=summary_rows)
    _append_material_summary_sheet(workbook, title="Bar-le-Duc", site="BAR_LE_DUC", rows=summary_rows)
    _append_material_detail_sheet(workbook, title="Detail Paris", site="PARIS", rows=detail_rows)
    _append_material_detail_sheet(workbook, title="Detail Bar-le-Duc", site="BAR_LE_DUC", rows=detail_rows)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _generated_report_out(row: GeneratedReport) -> GeneratedReportOut:
    return GeneratedReportOut(
        id=row.id,
        report_type=row.report_type,
        report_label=row.report_label,
        file_format=row.file_format,
        period_start=row.period_start,
        period_end=row.period_end,
        note=row.note,
        row_count=row.row_count,
        created_by_user_id=row.created_by_user_id,
        created_at=row.created_at,
    )


def _parse_report_date(value: object | None) -> date | None:
    raw = _text(value)
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _report_period_label(start: date | None, end: date | None) -> str:
    if start and end:
        return f"{start.isoformat()} - {end.isoformat()}"
    if start:
        return f"Depuis {start.isoformat()}"
    if end:
        return f"Jusqu au {end.isoformat()}"
    return "-"


def _decimal_from_mapping(value: object | None) -> dict[str, Decimal]:
    if not isinstance(value, dict):
        return {}
    out: dict[str, Decimal] = {}
    for currency, amount in value.items():
        currency_code = _text(currency).upper() or "EUR"
        try:
            out[currency_code] = Decimal(str(amount or "0")).quantize(Decimal("0.01"))
        except Exception:
            continue
    return out


def _money_label_from_mapping(value: object | None, fallback_currency: str = "EUR") -> str:
    amounts = _decimal_from_mapping(value)
    if not amounts:
        return f"0.00 {fallback_currency.upper()}"
    return " | ".join(f"{amount:.2f} {currency}" for currency, amount in sorted(amounts.items()))


def _sum_decimal_mapping(value: object | None) -> Decimal:
    total = Decimal("0.00")
    for amount in _decimal_from_mapping(value).values():
        total += amount
    return total.quantize(Decimal("0.01"))


def _money_label_from_scalar(value: object | None, currency: str = "EUR") -> str:
    try:
        amount = Decimal(str(value or "0")).quantize(Decimal("0.01"))
    except Exception:
        amount = Decimal("0.00")
    return f"{amount:.2f} {currency.upper()}"


def _form_label(config: TypeformFormConfig | None) -> str | None:
    if config is None:
        return None
    config_json = _json_object(config.configuration_json)
    return _text(config_json.get("label")) or config.source_code


def _intake_family_key(normalized: dict[str, object]) -> str | None:
    email = _text(normalized.get("parent_email")).casefold()
    if email:
        return f"email:{email}"
    phone = re.sub(r"\D+", "", _text(normalized.get("parent_phone")))
    if phone:
        return f"phone:{phone}"
    parent_name = _normalize_token(
        " ".join(
            part
            for part in [
                _text(normalized.get("parent_last_name")),
                _text(normalized.get("parent_first_name")),
            ]
            if part
        )
    )
    return f"name:{parent_name}" if parent_name else None


def _format_slot_preferences(preferences: list[object], *, include_location: bool = False) -> str | None:
    labels: list[str] = []
    for item in preferences:
        if not isinstance(item, dict):
            continue
        day = _text(item.get("day"))
        time = _text(item.get("time"))
        location = _text(item.get("location"))
        parts: list[str] = []
        if day and time:
            parts.append(f"{day.capitalize()} {time}")
        elif day:
            parts.append(day.capitalize())
        elif time:
            parts.append(time)
        if include_location and location:
            parts.append(location)
        label = " | ".join(parts).strip()
        if label:
            labels.append(label)
    return ", ".join(dict.fromkeys(labels)) or None


def _format_day_time_summary(normalized: dict[str, object]) -> str | None:
    days = [_text(item) for item in _json_list(normalized.get("requested_days")) if _text(item)]
    times = [_text(item) for item in _json_list(normalized.get("requested_times")) if _text(item)]
    labels: list[str] = []
    if days and times:
        for day in days:
            for time in times:
                labels.append(f"{day.capitalize()} {time}")
    elif days:
        labels.extend(day.capitalize() for day in days)
    else:
        labels.extend(times)
    return ", ".join(dict.fromkeys(labels)) or None


def _format_intake_course_1(normalized: dict[str, object]) -> str | None:
    location = _text(normalized.get("requested_location"))
    mode = _text(normalized.get("requested_course_mode"))
    formula = _text(normalized.get("requested_formula_type"))
    slots = _format_slot_preferences(_json_list(normalized.get("requested_slot_preferences"))) or _format_day_time_summary(normalized)
    parts = [part for part in [location, slots, mode or formula] if part]
    return " | ".join(parts) or None


def _format_intake_course_2(normalized: dict[str, object]) -> str | None:
    second_course = _json_object(normalized.get("requested_second_course"))
    if not _bool_or_default(second_course.get("requested"), False):
        return None
    value = _text(second_course.get("value")) or _text(second_course.get("label")) or "2e cours"
    modality = _text(second_course.get("modality"))
    slots = _format_slot_preferences(_json_list(second_course.get("slot_preferences")), include_location=True)
    parts = [part for part in [value, modality, slots] if part]
    return " | ".join(parts) or value


def _format_intake_solfege(normalized: dict[str, object]) -> str | None:
    products = [_text(item) for item in _json_list(normalized.get("requested_products")) if _text(item)]
    has_solfege = (
        _bool_or_default(normalized.get("requested_onsite_solfege"), False)
        or _bool_or_default(normalized.get("requested_online_solfege"), False)
        or any("solfege" in _normalize_token(item) for item in products)
        or bool(_json_list(normalized.get("requested_solfege_slot_preferences")))
    )
    if not has_solfege:
        return None
    level = _text(normalized.get("estimated_solfege_level"))
    modality = _text(normalized.get("requested_solfege_modality"))
    if modality == "online":
        modality = "en ligne"
    elif modality == "onsite":
        modality = "presentiel"
    slots = _format_slot_preferences(_json_list(normalized.get("requested_solfege_slot_preferences")), include_location=True)
    parts = [part for part in [f"Niveau {level}" if level else None, modality, slots] if part]
    return " | ".join(parts) or "Oui"


def _requested_product_contains(normalized: dict[str, object], *tokens: str) -> bool:
    normalized_tokens = tuple(_normalize_token(token) for token in tokens)
    for item in _json_list(normalized.get("requested_products")):
        product = _normalize_token(item)
        if all(token in product for token in normalized_tokens):
            return True
    return False


def _format_intake_masterclass(normalized: dict[str, object]) -> str | None:
    return "Oui" if _requested_product_contains(normalized, "masterclass") else None


def _format_intake_pass_recup(normalized: dict[str, object]) -> str | None:
    if _bool_or_default(normalized.get("requested_pass_recup"), False):
        return "Oui"
    return "Oui" if _requested_product_contains(normalized, "pass", "recup") else None


def _format_money(value: object | None, currency: str | None = "EUR") -> str:
    try:
        amount = Decimal(str(value or "0")).quantize(Decimal("0.01"))
    except Exception:
        amount = Decimal("0.00")
    return f"{amount:.2f} {(currency or 'EUR').upper()}"


def _quote_parent_contact_from_meta(meta: dict[str, object]) -> dict[str, str]:
    typeform_meta = _json_object(meta.get("typeform_intake"))
    normalized = _json_object(typeform_meta.get("normalized_payload"))
    parent_referent = _json_object(meta.get("parent_referent"))
    return {
        "name": _display_name(
            parent_referent.get("first_name") or normalized.get("parent_first_name"),
            parent_referent.get("last_name") or normalized.get("parent_last_name"),
            fallback="",
        ),
        "email": _text(parent_referent.get("email") or normalized.get("parent_email") or meta.get("recipient_email")),
        "phone": _text(parent_referent.get("phone") or normalized.get("parent_phone")),
    }


def _quote_family_key(quote: Quote, prospect: Prospect | None, parent: Prospect | None, client: User | None) -> str:
    meta = _json_object(quote.meta)
    contact = _quote_parent_contact_from_meta(meta)
    if parent is not None:
        return f"parent-prospect:{parent.id}"
    if prospect is not None and prospect.parent_prospect_id is not None:
        return f"parent-prospect:{prospect.parent_prospect_id}"
    if contact["email"]:
        return f"email:{contact['email'].casefold()}"
    if client is not None and client.email:
        return f"client-email:{client.email.casefold()}"
    if prospect is not None and prospect.email:
        return f"prospect-email:{prospect.email.casefold()}"
    return f"quote:{quote.id}"


def _quote_family_contact(
    quote: Quote,
    prospect: Prospect | None,
    parent: Prospect | None,
    client: User | None,
) -> tuple[str, str | None, str | None, str | None]:
    meta_contact = _quote_parent_contact_from_meta(_json_object(quote.meta))
    parent_name = _display_name(parent.first_name, parent.last_name, fallback="") if parent is not None else ""
    client_name = _display_name(client.first_name, client.last_name, fallback="") if client is not None else ""
    prospect_name = _display_name(prospect.first_name, prospect.last_name, fallback="") if prospect is not None else ""
    label = parent_name or meta_contact["name"] or client_name or prospect_name or "Famille sans contact"
    email = (
        (parent.email if parent is not None else None)
        or meta_contact["email"]
        or (client.email if client is not None else None)
        or (prospect.email if prospect is not None else None)
    )
    phone = (
        (parent.phone if parent is not None else None)
        or meta_contact["phone"]
        or (client.mobile_phone_1 if client is not None else None)
        or (client.phone if client is not None else None)
        or (prospect.phone if prospect is not None else None)
    )
    return label, label if label != "Famille sans contact" else None, email, phone


def _quote_student_name(quote: Quote, prospect: Prospect | None, client: User | None) -> str:
    meta = _json_object(quote.meta)
    duplicated_name = _text(meta.get("duplicated_for_child_name"))
    if duplicated_name:
        return duplicated_name
    typeform_meta = _json_object(meta.get("typeform_intake"))
    normalized = _json_object(typeform_meta.get("normalized_payload"))
    normalized_child = _display_name(
        normalized.get("child_first_name") or normalized.get("student_first_name"),
        normalized.get("child_last_name") or normalized.get("student_last_name"),
        fallback="",
    )
    if normalized_child:
        return normalized_child
    if prospect is not None:
        return _display_name(prospect.first_name, prospect.last_name, fallback=prospect.email)
    if client is not None:
        return _display_name(client.first_name, client.last_name, fallback=client.email)
    return "Eleve non renseigne"


def _quote_child_family_surname(child_name: object | None) -> str:
    raw = unicodedata.normalize("NFKD", _text(child_name))
    ascii_text = raw.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", " ", ascii_text.casefold()).strip()
    parts = [part for part in normalized.split() if part]
    return parts[-1] if len(parts) >= 2 else ""


def _quote_child_family_surname_label(child_name: object | None) -> str:
    parts = [part for part in _text(child_name).split() if part]
    return parts[-1] if len(parts) >= 2 else ""


def _quote_line_summary(lines: list[QuoteLine], *, categories: set[str], max_items: int = 8) -> str | None:
    labels: list[str] = []
    for line in lines:
        category = (line.line_category or "").strip().lower()
        line_type = (line.line_type or "").strip().lower()
        if category not in categories or line_type == "subtotal":
            continue
        qty = Decimal(line.quantity or 0).quantize(Decimal("0.01"))
        amount = Decimal(line.amount_ttc or 0).quantize(Decimal("0.01"))
        title = _text(line.title) or _text(line.code) or "Ligne"
        if qty != Decimal("1.00"):
            title = f"{title} x{qty.normalize()}"
        labels.append(f"{title} ({amount:.2f})")
    if not labels:
        return None
    visible = labels[:max_items]
    if len(labels) > max_items:
        visible.append(f"+{len(labels) - max_items} ligne(s)")
    return " | ".join(visible)


def _quote_line_search_text(line: QuoteLine) -> str:
    meta = _json_object(line.meta)
    return _normalize_token(
        " ".join(
            _text(value)
            for value in [
                line.title,
                line.description,
                line.code,
                line.line_category,
                line.line_type,
                line.master_item_type,
                meta.get("typeform_automatic_line"),
                meta.get("recommendation_key"),
            ]
        )
    )


def _quote_line_schedule_keys(line: QuoteLine) -> set[str]:
    keys: set[str] = set()
    activity_id = _text(line.activity_id)
    meta = _json_object(line.meta)
    automatic_key = _text(meta.get("typeform_automatic_line"))
    recommendation_key = _text(meta.get("recommendation_key"))
    if activity_id:
        keys.add(activity_id)
    if automatic_key:
        keys.add(automatic_key)
    if recommendation_key:
        keys.add(recommendation_key)
    if activity_id and automatic_key:
        keys.add(f"{activity_id}:{automatic_key}")
    return {key for key in keys if key}


def _quote_line_strong_schedule_keys(line: QuoteLine) -> set[str]:
    activity_id = _text(line.activity_id)
    meta = _json_object(line.meta)
    automatic_key = _text(meta.get("typeform_automatic_line"))
    recommendation_key = _text(meta.get("recommendation_key"))
    keys = {key for key in [automatic_key, recommendation_key] if key}
    if activity_id and automatic_key:
        keys.add(f"{activity_id}:{automatic_key}")
    return keys


def _quote_schedule_item_keys(item: dict[str, object]) -> set[str]:
    keys: set[str] = set()
    for key in ("recommendation_key", "line_recommendation_key", "activity_id", "source_key", "typeform_automatic_line"):
        value = _text(item.get(key))
        if value:
            keys.add(value)
    activity_id = _text(item.get("activity_id"))
    source_key = _text(item.get("source_key") or item.get("typeform_automatic_line"))
    if activity_id and source_key:
        keys.add(f"{activity_id}:{source_key}")
    return keys


def _quote_schedule_items(quote: Quote) -> list[dict[str, object]]:
    snapshot = _json_object(quote.calendar_snapshot)
    items: list[dict[str, object]] = []
    for collection_key in ("blocks", "sessions"):
        for raw in _json_list(snapshot.get(collection_key)):
            item = _json_object(raw)
            if item:
                items.append(item)
    selected_slot = _json_object(quote.selected_solfege_slot) or _json_object(
        _json_object(snapshot.get("solfege")).get("selected_slot")
    )
    if selected_slot:
        slot_item = dict(selected_slot)
        slot_item.setdefault("activity_label", "Solfege")
        slot_item.setdefault("source", "selected_solfege_slot")
        slot_item.setdefault("pending_solfege_level", quote.estimated_solfege_level)
        items.append(slot_item)
    return items


def _quote_schedule_item_label(
    item: dict[str, object],
    *,
    include_level: bool = False,
    fallback_title: str | None = None,
) -> str | None:
    day = _text(item.get("weekday_label") or item.get("day") or item.get("date_label"))
    start = _text(item.get("start_time") or item.get("start") or item.get("local_start_time"))
    end = _text(item.get("end_time") or item.get("end") or item.get("local_end_time"))
    if not day:
        starts_at = _text(item.get("start_at") or item.get("start_at_local") or item.get("start_at_utc") or item.get("date"))
        day = starts_at[:10] if starts_at else ""
    time_label = f"{start}-{end}" if start and end else start or end
    location = _text(
        item.get("location_label")
        or item.get("location_name")
        or item.get("location")
        or item.get("modality_label")
        or item.get("modality")
    )
    level = _text(item.get("pending_solfege_level") or item.get("level_code") or item.get("level"))
    parts = []
    if include_level and level:
        parts.append(f"Niveau {level}")
    parts.extend(part for part in [day, time_label, location] if part)
    if parts:
        return " · ".join(parts)
    label = _text(item.get("label"))
    if label:
        return label
    return fallback_title


def _quote_schedule_item_is_kind(item: dict[str, object], token: str) -> bool:
    item_text = _normalize_token(
        " ".join(
            _text(item.get(key))
            for key in (
                "activity_label",
                "activity_name",
                "course_type_name",
                "title",
                "label",
                "activity_code",
                "activity_service_code",
            )
        )
    )
    return token in item_text


def _quote_primary_course_schedule_items(quote: Quote) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for item in _quote_schedule_items(quote):
        if _quote_schedule_item_is_kind(item, "solfege"):
            continue
        if _quote_schedule_item_is_kind(item, "masterclass"):
            continue
        if _text(item.get("pending_solfege_level")):
            continue
        label = _quote_schedule_item_label(item)
        if label:
            items.append(item)
    return items


def _quote_schedule_label_has_slot(label: str | None) -> bool:
    if not label:
        return False
    return bool(re.search(r"\d{1,2}:\d{2}", label)) and " · " in label


def _quote_line_schedule_label(quote: Quote, line: QuoteLine | None, *, include_level: bool = False) -> str | None:
    if line is None:
        return None
    strong_line_keys = _quote_line_strong_schedule_keys(line)
    activity_id = _text(line.activity_id)
    line_text = _quote_line_search_text(line)
    fallback_title = _text(line.title or line.description or line.code) or None
    items = _quote_schedule_items(quote)
    for item in items:
        if strong_line_keys and strong_line_keys.intersection(_quote_schedule_item_keys(item)):
            return _quote_schedule_item_label(item, include_level=include_level, fallback_title=fallback_title)
    if activity_id:
        activity_matches = [item for item in items if activity_id in _quote_schedule_item_keys(item)]
        if len(activity_matches) == 1:
            return _quote_schedule_item_label(activity_matches[0], include_level=include_level, fallback_title=fallback_title)
    for item in items:
        if "solfege" in line_text and _quote_schedule_item_is_kind(item, "solfege"):
            return _quote_schedule_item_label(item, include_level=include_level, fallback_title=fallback_title)
        if "masterclass" in line_text and _quote_schedule_item_is_kind(item, "masterclass"):
            return _quote_schedule_item_label(item, include_level=include_level, fallback_title=fallback_title)
    return fallback_title


def _quote_family_child_schedule(quote: Quote, lines: list[QuoteLine]) -> dict[str, str | None]:
    service_lines: list[QuoteLine] = []
    solfege_line: QuoteLine | None = None
    masterclass_line: QuoteLine | None = None
    second_course_line: QuoteLine | None = None

    for line in lines:
        line_type = (line.line_type or "").strip().lower()
        if line_type == "subtotal":
            continue
        text = _quote_line_search_text(line)
        if "masterclass" in text:
            masterclass_line = masterclass_line or line
            continue
        if "solfege" in text:
            solfege_line = solfege_line or line
            continue
        is_service = (line.line_category or "").strip().lower() == "service" or line.activity_id is not None
        if not is_service:
            continue
        if any(token in text for token in ("second", "deuxieme", "2e", "2eme", "secondaire")):
            second_course_line = second_course_line or line
        service_lines.append(line)

    course_1_line = next((line for line in service_lines if line is not second_course_line), None)
    if second_course_line is None and len(service_lines) >= 2:
        second_course_line = service_lines[1]

    course_items = _quote_primary_course_schedule_items(quote)
    course_1 = _quote_line_schedule_label(quote, course_1_line)
    course_2 = _quote_line_schedule_label(quote, second_course_line)
    if not _quote_schedule_label_has_slot(course_1) and course_items:
        course_1 = _quote_schedule_item_label(course_items[0], fallback_title=course_1)
    if not _quote_schedule_label_has_slot(course_2) and len(course_items) >= 2:
        course_2 = _quote_schedule_item_label(course_items[1], fallback_title=course_2)

    return {
        "course_1": course_1,
        "course_2": course_2,
        "solfege": _quote_line_schedule_label(quote, solfege_line, include_level=True),
        "masterclass": _quote_line_schedule_label(quote, masterclass_line),
    }


def _quote_planning_summary(quote: Quote, max_items: int = 8) -> str | None:
    snapshot = _json_object(quote.calendar_snapshot)
    sessions = _json_list(snapshot.get("sessions"))
    labels: list[str] = []
    for raw in sessions:
        item = _json_object(raw)
        title = _text(item.get("course_type_name") or item.get("activity_name") or item.get("title"))
        location = _text(item.get("location_name") or item.get("location"))
        day = _text(item.get("day") or item.get("weekday_label"))
        start = _text(item.get("start_time") or item.get("start") or item.get("local_start_time"))
        end = _text(item.get("end_time") or item.get("end") or item.get("local_end_time"))
        if not day:
            starts_at = _text(item.get("start_at") or item.get("start_at_local") or item.get("start_at_utc"))
            day = starts_at[:10] if starts_at else ""
        time_label = f"{start}-{end}" if start and end else start or end
        parts = [part for part in [title, location, day, time_label] if part]
        if parts:
            labels.append(" · ".join(parts))
    if not labels:
        return None
    visible = list(dict.fromkeys(labels))[:max_items]
    if len(labels) > max_items:
        visible.append(f"+{len(labels) - max_items} seance(s)")
    return " | ".join(visible)


def _quote_payment_summary(quote: Quote) -> str | None:
    snapshot = _json_object(quote.payment_terms_snapshot)
    plan_label = _text(snapshot.get("payment_plan_name") or snapshot.get("plan_name") or snapshot.get("payment_method_label"))
    schedule = _json_list(snapshot.get("schedule"))
    schedule_labels: list[str] = []
    for raw in schedule[:4]:
        item = _json_object(raw)
        amount = _text(item.get("amount_ttc") or item.get("amount") or item.get("total_ttc"))
        due = _text(item.get("due_date") or item.get("label") or item.get("deposit_label"))
        if amount or due:
            schedule_labels.append(" ".join(part for part in [due, amount] if part))
    parts = [plan_label, " | ".join(schedule_labels)]
    return " ; ".join(part for part in parts if part) or None


def _merge_quote_family_groups_by_child_surname(
    grouped: dict[str, dict[str, object]],
    latest_created: dict[str, datetime],
) -> tuple[dict[str, dict[str, object]], dict[str, datetime]]:
    buckets_by_surname: dict[tuple[str, str], list[dict[str, object]]] = {}
    for family in grouped.values():
        surnames = {
            _quote_child_family_surname(_json_object(child).get("child_name"))
            for child in _json_list(family.get("children"))
        }
        surnames.discard("")
        family_label_key = _normalize_token(family.get("family_label"))
        if len(surnames) == 1 and family_label_key and family_label_key != "famillesanscontact":
            buckets_by_surname.setdefault((next(iter(surnames)), family_label_key), []).append(family)

    merged_keys: set[str] = set()
    merged_grouped: dict[str, dict[str, object]] = {}
    merged_latest: dict[str, datetime] = {}
    for (surname, _), families in buckets_by_surname.items():
        if len(families) <= 1:
            continue
        merged_children: list[object] = []
        contacts_email: list[str] = []
        contacts_phone: list[str] = []
        surname_label = ""
        quote_count = 0
        latest = datetime.min.replace(tzinfo=timezone.utc)
        for family in families:
            family_key = _text(family.get("family_key"))
            merged_keys.add(family_key)
            latest = max(latest, latest_created.get(family_key, latest))
            quote_count += int(family.get("quote_count") or 0)
            for child in _json_list(family.get("children")):
                child_obj = _json_object(child)
                surname_label = surname_label or _quote_child_family_surname_label(child_obj.get("child_name"))
                merged_children.append(child)
            for key, target in (("parent_email", contacts_email), ("parent_phone", contacts_phone)):
                value = _text(family.get(key))
                if value and value not in target:
                    target.append(value)
        merged_key = f"child-surname:{surname}"
        merged_grouped[merged_key] = {
            "family_key": merged_key,
            "family_label": f"Famille {surname_label or surname.title()}",
            "parent_name": None,
            "parent_email": " / ".join(contacts_email[:3]) if contacts_email else None,
            "parent_phone": " / ".join(contacts_phone[:3]) if contacts_phone else None,
            "quote_count": quote_count,
            "children": merged_children,
        }
        merged_latest[merged_key] = latest

    for family_key, family in grouped.items():
        if family_key in merged_keys:
            continue
        merged_grouped[family_key] = family
        merged_latest[family_key] = latest_created.get(family_key, datetime.min.replace(tzinfo=timezone.utc))
    return merged_grouped, merged_latest


def _build_quote_family_summary_rows(
    db: Session,
    *,
    q: str | None = None,
    school_year_label: str | None = None,
    received_from: date | None = None,
    received_to: date | None = None,
    status_filter: str | None = None,
    min_children: int = 2,
    limit: int = 1000,
) -> list[dict[str, object]]:
    if received_from is not None and received_to is not None and received_from > received_to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'received_from' must be before 'received_to'",
        )

    stmt = (
        select(Quote, Prospect, User)
        .outerjoin(Prospect, Prospect.id == Quote.prospect_id)
        .outerjoin(User, User.id == Quote.client_id)
        .order_by(Quote.created_at.desc())
        .limit(limit)
    )
    if school_year_label:
        stmt = stmt.where(Quote.school_year_label == school_year_label)
    if received_from is not None:
        start_local, _ = _day_bounds(received_from)
        stmt = stmt.where(Quote.created_at >= start_local)
    if received_to is not None:
        _, end_local = _day_bounds(received_to)
        stmt = stmt.where(Quote.created_at < end_local)
    if status_filter:
        stmt = stmt.where(Quote.status.ilike(status_filter.strip()))
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Quote.quote_number.ilike(like),
                Quote.status.ilike(like),
                Quote.school_year_label.ilike(like),
                cast(Quote.meta, Text).ilike(like),
                Prospect.first_name.ilike(like),
                Prospect.last_name.ilike(like),
                Prospect.email.ilike(like),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
                User.email.ilike(like),
            )
        )

    rows = db.execute(stmt).all()
    parent_ids = {
        prospect.parent_prospect_id
        for _, prospect, _ in rows
        if prospect is not None and prospect.parent_prospect_id is not None
    }
    parents_by_id: dict[UUID, Prospect] = {}
    if parent_ids:
        parents_by_id = {
            parent.id: parent
            for parent in db.scalars(select(Prospect).where(Prospect.id.in_(list(parent_ids)))).all()
        }
    quote_ids = [quote.id for quote, _, _ in rows]
    lines_by_quote_id: dict[UUID, list[QuoteLine]] = {quote_id: [] for quote_id in quote_ids}
    if quote_ids:
        for line in db.scalars(
            select(QuoteLine)
            .where(QuoteLine.quote_id.in_(quote_ids))
            .order_by(QuoteLine.quote_id.asc(), QuoteLine.sort_order.asc(), QuoteLine.created_at.asc())
        ).all():
            lines_by_quote_id.setdefault(line.quote_id, []).append(line)

    grouped: dict[str, dict[str, object]] = {}
    latest_created: dict[str, datetime] = {}
    for quote, prospect, client in rows:
        parent = parents_by_id.get(prospect.parent_prospect_id) if prospect is not None and prospect.parent_prospect_id is not None else None
        family_key = _quote_family_key(quote, prospect, parent, client)
        family_label, parent_name, parent_email, parent_phone = _quote_family_contact(quote, prospect, parent, client)
        bucket = grouped.get(family_key)
        if bucket is None:
            bucket = {
                "family_key": family_key,
                "family_label": family_label,
                "parent_name": parent_name,
                "parent_email": parent_email,
                "parent_phone": parent_phone,
                "quote_count": 0,
                "children": [],
            }
            grouped[family_key] = bucket
            latest_created[family_key] = quote.created_at
        else:
            bucket["parent_name"] = bucket.get("parent_name") or parent_name
            bucket["parent_email"] = bucket.get("parent_email") or parent_email
            bucket["parent_phone"] = bucket.get("parent_phone") or parent_phone
            if bucket.get("family_label") == "Famille sans contact" and family_label != "Famille sans contact":
                bucket["family_label"] = family_label
            if quote.created_at > latest_created[family_key]:
                latest_created[family_key] = quote.created_at

        lines = lines_by_quote_id.get(quote.id, [])
        child_schedule = _quote_family_child_schedule(quote, lines)
        bucket["quote_count"] = int(bucket.get("quote_count") or 0) + 1
        _json_list(bucket["children"]).append(
            {
                "quote_id": str(quote.id),
                "quote_number": quote.quote_number,
                "created_at": quote.created_at.isoformat(),
                "sent_at": quote.sent_at.isoformat() if quote.sent_at else None,
                "expires_at": quote.expires_at.isoformat() if quote.expires_at else None,
                "child_name": _quote_student_name(quote, prospect, client),
                "status": quote.status,
                "total_ttc": _format_money(quote.total_ttc, quote.currency),
                "course_1": child_schedule.get("course_1"),
                "course_2": child_schedule.get("course_2"),
                "solfege": child_schedule.get("solfege"),
                "masterclass": child_schedule.get("masterclass"),
                "services": _quote_line_summary(lines, categories={"service"}),
                "products": _quote_line_summary(lines, categories={"product", "kit"}),
                "discounts": _quote_line_summary(lines, categories={"discount", "adjustment", "fee"}),
                "planning": _quote_planning_summary(quote),
                "payment": _quote_payment_summary(quote),
            }
        )

    grouped, latest_created = _merge_quote_family_groups_by_child_surname(grouped, latest_created)
    families = [
        family
        for family in grouped.values()
        if len({_text(_json_object(child).get("child_name")).casefold() for child in _json_list(family.get("children"))}) >= min_children
    ]
    for family in families:
        _json_list(family.get("children")).sort(
            key=lambda child: (
                _text(_json_object(child).get("child_name")).casefold(),
                _text(_json_object(child).get("quote_number")),
            )
        )
    families.sort(
        key=lambda family: (
            -len(_json_list(family.get("children"))),
            _text(family.get("family_label")).casefold(),
            latest_created.get(_text(family.get("family_key")), datetime.min.replace(tzinfo=timezone.utc)),
        )
    )
    return families


def _build_expired_quote_rows(
    db: Session,
    *,
    q: str | None = None,
    school_year_label: str | None = None,
    expired_from: date | None = None,
    expired_to: date | None = None,
    status_filter: str | None = None,
    limit: int = 5000,
) -> list[dict[str, object]]:
    if expired_from is not None and expired_to is not None and expired_from > expired_to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'expired_from' must be before 'expired_to'",
        )

    now_utc = datetime.now(timezone.utc)
    terminal_expired = or_(
        Quote.status == "expired",
        (
            Quote.status.in_(["sent", "change_requested"])
            & Quote.expires_at.is_not(None)
            & (Quote.expires_at < now_utc)
        ),
    )
    effective_status_expr = case(
        (Quote.status == "rejected", "rejected"),
        (Quote.status == "cancelled", "cancelled"),
        (terminal_expired, "expired"),
        else_=None,
    )
    effective_date_expr = case(
        (Quote.status == "rejected", func.coalesce(Quote.rejected_at, Quote.updated_at)),
        (Quote.status == "cancelled", func.coalesce(Quote.cancelled_at, Quote.updated_at)),
        (terminal_expired, func.coalesce(Quote.expired_at, Quote.expires_at)),
        else_=None,
    )

    stmt = (
        select(
            Quote,
            Prospect,
            User,
            Location,
            effective_status_expr.label("effective_status"),
            effective_date_expr.label("effective_date"),
        )
        .outerjoin(Prospect, Prospect.id == Quote.prospect_id)
        .outerjoin(User, User.id == Quote.client_id)
        .outerjoin(Location, Location.id == Quote.location_id)
        .where(effective_status_expr.is_not(None), effective_date_expr.is_not(None))
        .order_by(effective_date_expr.desc(), Quote.quote_number.asc())
        .limit(limit)
    )
    if school_year_label:
        stmt = stmt.where(Quote.school_year_label == school_year_label)
    if expired_from is not None:
        start_local, _ = _day_bounds(expired_from)
        stmt = stmt.where(effective_date_expr >= start_local)
    if expired_to is not None:
        _, end_local = _day_bounds(expired_to)
        stmt = stmt.where(effective_date_expr < end_local)
    if status_filter:
        normalized_status = status_filter.strip().casefold()
        status_aliases = {
            "expire": "expired",
            "expired": "expired",
            "expires": "expired",
            "refuse": "rejected",
            "refused": "rejected",
            "rejected": "rejected",
            "annule": "cancelled",
            "cancelled": "cancelled",
            "canceled": "cancelled",
        }
        stmt = stmt.where(effective_status_expr == status_aliases.get(normalized_status, normalized_status))
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Quote.quote_number.ilike(like),
                Quote.status.ilike(like),
                Quote.school_year_label.ilike(like),
                cast(Quote.meta, Text).ilike(like),
                Prospect.first_name.ilike(like),
                Prospect.last_name.ilike(like),
                Prospect.email.ilike(like),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
                User.email.ilike(like),
            )
        )

    rows = db.execute(stmt).all()
    quote_ids = [quote.id for quote, _, _, _, _, _ in rows]
    lines_by_quote_id: dict[UUID, list[QuoteLine]] = {quote_id: [] for quote_id in quote_ids}
    if quote_ids:
        for line in db.scalars(
            select(QuoteLine)
            .where(QuoteLine.quote_id.in_(quote_ids))
            .order_by(QuoteLine.quote_id.asc(), QuoteLine.sort_order.asc(), QuoteLine.created_at.asc())
        ).all():
            lines_by_quote_id.setdefault(line.quote_id, []).append(line)

    status_labels = {
        "expired": "Expire",
        "rejected": "Refuse par le client",
        "cancelled": "Annule par l admin",
    }
    out: list[dict[str, object]] = []
    for quote, prospect, client, location, effective_status, effective_date in rows:
        contact = _quote_parent_contact_from_meta(_json_object(quote.meta))
        contact_name = contact["name"] or (
            _display_name(client.first_name, client.last_name, fallback="") if client is not None else ""
        ) or (
            _display_name(prospect.first_name, prospect.last_name, fallback="") if prospect is not None else ""
        )
        contact_email = contact["email"] or (client.email if client is not None else "") or (prospect.email if prospect is not None else "")
        contact_phone = contact["phone"] or (client.mobile_phone_1 if client is not None else "") or (client.phone if client is not None else "") or (
            prospect.phone if prospect is not None else ""
        )
        lines = lines_by_quote_id.get(quote.id, [])
        out.append(
            {
                "quote_id": str(quote.id),
                "quote_number": quote.quote_number,
                "student_name": _quote_student_name(quote, prospect, client),
                "contact_name": contact_name,
                "contact_email": contact_email,
                "contact_phone": contact_phone,
                "status": quote.status,
                "effective_status": effective_status,
                "effective_status_label": status_labels.get(str(effective_status), str(effective_status or quote.status)),
                "effective_date": effective_date.isoformat() if effective_date else None,
                "created_at": quote.created_at.isoformat(),
                "sent_at": quote.sent_at.isoformat() if quote.sent_at else None,
                "expires_at": quote.expires_at.isoformat() if quote.expires_at else None,
                "expired_at": quote.expired_at.isoformat() if quote.expired_at else None,
                "cancelled_at": quote.cancelled_at.isoformat() if quote.cancelled_at else None,
                "rejected_at": quote.rejected_at.isoformat() if quote.rejected_at else None,
                "school_year_label": quote.school_year_label,
                "location": location.name if location is not None else None,
                "total_ttc": _format_money(quote.total_ttc, quote.currency),
                "planning": _quote_planning_summary(quote),
                "services": _quote_line_summary(lines, categories={"service"}),
                "products": _quote_line_summary(lines, categories={"product", "kit"}),
                "payment": _quote_payment_summary(quote),
            }
        )
    return out


def _build_overdue_invoice_rows(
    db: Session,
    *,
    q: str | None = None,
    due_from: date | None = None,
    due_to: date | None = None,
    limit: int = 5000,
) -> list[dict[str, object]]:
    if due_from is not None and due_to is not None and due_from > due_to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'due_from' must be before 'due_to'",
        )

    today = datetime.now(ADMIN_COMMUNICATION_TIMEZONE).date()
    effective_due_to = due_to or today
    search_token = q.strip().casefold() if q else ""
    rows = db.execute(
        select(ClientNoteEntry, User)
        .join(User, User.id == ClientNoteEntry.user_id)
        .where(ClientNoteEntry.message.contains(INVOICE_RANGE_NOTE_PREFIX))
        .order_by(ClientNoteEntry.created_at.desc())
        .limit(limit)
    ).all()

    out: list[dict[str, object]] = []
    for note, client in rows:
        metadata = _invoice_range_metadata_from_note_message(note.message)
        if metadata is None:
            continue
        invoice_status = _text(metadata.get("invoice_status")).upper() or "ISSUED"
        if invoice_status != "ISSUED":
            continue
        if _bool_or_default(metadata.get("no_due_date"), False):
            continue
        due_date = _parse_report_date(metadata.get("due_date"))
        if due_date is None:
            continue
        if due_from is not None and due_date < due_from:
            continue
        if due_date > effective_due_to:
            continue

        invoice_number = _text(metadata.get("invoice_number")) or "-"
        client_name = _text(metadata.get("client_name")) or _client_name(client)
        client_email = _text(client.email)
        client_phone = _text(client.mobile_phone_1) or _text(client.mobile_phone_2) or _text(client.phone) or _text(client.home_phone)
        billing_entity = _text(metadata.get("billing_entity")) or "-"
        amounts = metadata.get("total_to_pay_by_currency") or metadata.get("totals_by_currency")
        amount_currency = next(iter(_decimal_from_mapping(amounts).keys()), _text(metadata.get("currency")) or "EUR")
        paid_value = metadata.get("payment_amount_paid")
        paid_label = _money_label_from_scalar(paid_value, amount_currency) if paid_value not in (None, "") else "-"
        emailed_at = _text(metadata.get("emailed_at"))[:10]
        reminded_at = _text(metadata.get("reminded_at"))[:10]

        searchable = " ".join(
            [
                invoice_number,
                client_name,
                client_email,
                client_phone,
                billing_entity,
                _text(metadata.get("private_note")),
                str(note.id),
            ]
        ).casefold()
        if search_token and search_token not in searchable:
            continue

        out.append(
            {
                "note_id": str(note.id),
                "client_id": str(client.id),
                "invoice_number": invoice_number,
                "client_name": client_name,
                "client_email": client_email,
                "client_phone": client_phone,
                "billing_entity": billing_entity,
                "issued_date": _text(metadata.get("issued_date"))[:10] or note.created_at.date().isoformat(),
                "due_date": due_date.isoformat(),
                "days_overdue": max(0, (today - due_date).days),
                "total_due": _money_label_from_mapping(amounts, amount_currency),
                "amount_paid": paid_label,
                "status": "Emise non payee",
                "emailed_at": emailed_at,
                "reminded_at": reminded_at,
                "last_contact_at": reminded_at or emailed_at or "-",
            }
        )
    out.sort(key=lambda item: (_text(item.get("due_date")), _text(item.get("client_name")).casefold(), _text(item.get("invoice_number"))))
    return out


@router.post("/communications/{communication_id}/resend", response_model=CommunicationReportRow)
def resend_communication(
    communication_id: str = Path(...),
    payload: CommunicationResendRequest | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> CommunicationReportRow:
    normalized_communication_id = _parse_communication_log_id(communication_id)
    row = db.scalar(select(CommunicationLog).where(CommunicationLog.id == normalized_communication_id).limit(1))
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Communication introuvable")
    if row.channel != CommunicationChannelModel.EMAIL:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Seules les communications email peuvent etre renvoyees")

    delivery_error = email_delivery_disabled_reason()
    if delivery_error:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=delivery_error)

    recipient = str((payload.recipient_email if payload is not None else None) or row.recipient or "").strip().lower()
    if not recipient:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Destinataire email introuvable")
    sender_category = getattr(row.sender_category, "value", str(row.sender_category or "")).strip().upper()
    sender = resolve_sender_profile(db, sender_kind="TEACHER" if sender_category == "PROFESSOR" else "STUDIO")

    message_id = send_email(
        to_email=recipient,
        subject=row.subject,
        body=row.content,
        body_format=row.content_format.value if hasattr(row.content_format, "value") else str(row.content_format),
        context=f"{row.source}_RESEND",
        from_email=sender.from_email,
        from_name=sender.from_name,
        reply_to=sender.reply_to,
        subject_prefix=sender.subject_prefix,
        sender_user_id=row.sender_user_id,
        sender_label=row.sender_label,
        sender_category=row.sender_category,
        professor_id=row.professor_id,
        recipient_user_id=row.recipient_user_id if recipient == str(row.recipient or "").strip().lower() else None,
        communication_type=row.communication_type,
    )
    resent_row = db.scalar(
        select(CommunicationLog)
        .where(
            CommunicationLog.provider_message_id == message_id,
            CommunicationLog.channel == CommunicationChannelModel.EMAIL,
        )
        .order_by(CommunicationLog.occurred_at.desc())
        .limit(1)
    )
    if resent_row is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Renvoi journalise introuvable")
    return _communication_row_out(resent_row)


@router.get("/reservations", response_model=list[ReservationReportRow])
def report_reservations(
    from_: datetime | None = Query(default=None, alias="from"),
    to: datetime | None = None,
    course_type_id: UUID | None = None,
    location_id: UUID | None = None,
    professor_id: UUID | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[ReservationReportRow]:
    _ensure_date_range(from_, to)

    stmt = (
        select(CourseSession, CourseType, Location, Professor, Booking, User)
        .join(CourseType, CourseType.id == CourseSession.course_type_id)
        .join(Location, Location.id == CourseSession.location_id)
        .join(Professor, Professor.id == CourseSession.professor_id)
        .join(Booking, Booking.session_id == CourseSession.id)
        .join(User, User.id == Booking.user_id)
    )

    if from_ is not None:
        stmt = stmt.where(CourseSession.start_at_utc >= from_)
    if to is not None:
        stmt = stmt.where(CourseSession.start_at_utc <= to)
    if course_type_id is not None:
        stmt = stmt.where(CourseSession.course_type_id == course_type_id)
    if location_id is not None:
        stmt = stmt.where(CourseSession.location_id == location_id)
    if professor_id is not None:
        stmt = stmt.where(CourseSession.professor_id == professor_id)

    rows = db.execute(stmt.order_by(CourseSession.start_at_utc.desc(), Booking.booked_at.desc())).all()

    return [
        ReservationReportRow(
            session_id=session.id,
            start_at_utc=session.start_at_utc,
            end_at_utc=session.end_at_utc,
            session_status=session.status,
            course_type_id=course_type.id,
            course_type_code=course_type.code,
            course_type_name=course_type.name,
            location_id=location.id,
            location_name=location.name,
            professor_id=professor.id,
            professor_name=_professor_name(professor),
            booking_id=booking.id,
            client_email=user.email,
            booking_status=booking.status,
            total_incl_vat_snapshot=booking.total_incl_vat_snapshot,
            currency_snapshot=booking.currency_snapshot,
        )
        for session, course_type, location, professor, booking, user in rows
    ]


@router.get("/attendance", response_model=list[AttendanceReportRow])
def report_attendance(
    from_: datetime | None = Query(default=None, alias="from"),
    to: datetime | None = None,
    course_type_id: UUID | None = None,
    location_id: UUID | None = None,
    professor_id: UUID | None = None,
    include_cancelled: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[AttendanceReportRow]:
    _ensure_date_range(from_, to)

    stmt = (
        select(CourseSession, CourseType, Location, Professor, Booking, User)
        .join(CourseType, CourseType.id == CourseSession.course_type_id)
        .join(Location, Location.id == CourseSession.location_id)
        .join(Professor, Professor.id == CourseSession.professor_id)
        .join(Booking, Booking.session_id == CourseSession.id)
        .join(User, User.id == Booking.user_id)
    )

    if from_ is not None:
        stmt = stmt.where(CourseSession.start_at_utc >= from_)
    if to is not None:
        stmt = stmt.where(CourseSession.start_at_utc <= to)
    if course_type_id is not None:
        stmt = stmt.where(CourseSession.course_type_id == course_type_id)
    if location_id is not None:
        stmt = stmt.where(CourseSession.location_id == location_id)
    if professor_id is not None:
        stmt = stmt.where(CourseSession.professor_id == professor_id)

    stmt = stmt.where(Booking.status != BookingStatus.WAITLISTED)
    if not include_cancelled:
        stmt = stmt.where(Booking.status != BookingStatus.CANCELLED)

    rows = db.execute(stmt.order_by(CourseSession.start_at_utc.desc(), Booking.booked_at.desc())).all()

    return [
        AttendanceReportRow(
            session_id=session.id,
            start_at_utc=session.start_at_utc,
            course_type_name=course_type.name,
            location_name=location.name,
            professor_name=_professor_name(professor),
            booking_id=booking.id,
            client_email=user.email,
            attendance_status=booking.status.value,
        )
        for session, course_type, location, professor, booking, user in rows
    ]


@router.get("/professor-statements", response_model=list[ProfessorStatementRow])
def report_professor_statements(
    from_: datetime | None = Query(default=None, alias="from"),
    to: datetime | None = None,
    professor_id: UUID | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[ProfessorStatementRow]:
    _ensure_date_range(from_, to)

    duration_hours_expr = cast(
        extract("epoch", CourseSession.end_at_utc - CourseSession.start_at_utc) / 3600.0,
        Numeric(10, 2),
    )

    booked_case = case(
        (
            Booking.status.in_(
                [
                    BookingStatus.BOOKED,
                    BookingStatus.ATTENDED,
                    BookingStatus.NO_SHOW,
                    BookingStatus.EXCUSED_ABSENCE,
                ]
            ),
            1,
        ),
        else_=0,
    )
    attended_case = case((Booking.status == BookingStatus.ATTENDED, 1), else_=0)
    no_show_case = case((Booking.status == BookingStatus.NO_SHOW, 1), else_=0)
    excused_case = case((Booking.status == BookingStatus.EXCUSED_ABSENCE, 1), else_=0)

    stmt = (
        select(
            CourseSession.id.label("session_id"),
            Professor.id.label("professor_id"),
            Professor.first_name.label("prof_first_name"),
            Professor.last_name.label("prof_last_name"),
            CourseSession.start_at_utc,
            CourseSession.end_at_utc,
            CourseSession.status,
            CourseType.name.label("course_type_name"),
            Location.name.label("location_name"),
            duration_hours_expr.label("duration_hours"),
            func.coalesce(func.sum(booked_case), 0).label("booked_students"),
            func.coalesce(func.sum(attended_case), 0).label("attended_students"),
            func.coalesce(func.sum(no_show_case), 0).label("no_show_students"),
            func.coalesce(func.sum(excused_case), 0).label("excused_absence_students"),
            ProfessorSessionPayout.hourly_rate_snapshot.label("hourly_rate_snapshot"),
            ProfessorSessionPayout.amount_snapshot.label("amount_snapshot"),
            ProfessorSessionPayout.currency_snapshot.label("currency_snapshot"),
            ProfessorSessionPayout.payout_status.label("payout_status"),
        )
        .join(Professor, Professor.id == CourseSession.professor_id)
        .join(CourseType, CourseType.id == CourseSession.course_type_id)
        .join(Location, Location.id == CourseSession.location_id)
        .outerjoin(Booking, Booking.session_id == CourseSession.id)
        .outerjoin(ProfessorSessionPayout, ProfessorSessionPayout.session_id == CourseSession.id)
        .where(CourseSession.status != SessionStatus.CANCELLED)
        .group_by(
            CourseSession.id,
            Professor.id,
            Professor.first_name,
            Professor.last_name,
            CourseSession.start_at_utc,
            CourseSession.end_at_utc,
            CourseSession.status,
            CourseType.name,
            Location.name,
            duration_hours_expr,
            ProfessorSessionPayout.hourly_rate_snapshot,
            ProfessorSessionPayout.amount_snapshot,
            ProfessorSessionPayout.currency_snapshot,
            ProfessorSessionPayout.payout_status,
        )
    )

    if from_ is not None:
        stmt = stmt.where(CourseSession.start_at_utc >= from_)
    if to is not None:
        stmt = stmt.where(CourseSession.start_at_utc <= to)
    if professor_id is not None:
        stmt = stmt.where(CourseSession.professor_id == professor_id)

    rows = db.execute(stmt.order_by(CourseSession.start_at_utc.desc())).all()

    result: list[ProfessorStatementRow] = []
    for row in rows:
        result.append(
            ProfessorStatementRow(
                session_id=row.session_id,
                professor_id=row.professor_id,
                professor_name=f"{row.prof_first_name} {row.prof_last_name}".strip(),
                start_at_utc=row.start_at_utc,
                end_at_utc=row.end_at_utc,
                session_status=row.status,
                course_type_name=row.course_type_name,
                location_name=row.location_name,
                duration_hours=float(row.duration_hours or 0),
                booked_students=int(row.booked_students or 0),
                attended_students=int(row.attended_students or 0),
                no_show_students=int(row.no_show_students or 0),
                excused_absence_students=int(row.excused_absence_students or 0),
                hourly_rate_snapshot=row.hourly_rate_snapshot,
                amount_snapshot=row.amount_snapshot,
                currency_snapshot=row.currency_snapshot,
                payout_status=(row.payout_status.value if row.payout_status is not None else None),
            )
        )

    return result


def _build_intake_family_summary_rows(
    db: Session,
    *,
    q: str | None = None,
    school_year_label: str | None = None,
    received_from: date | None = None,
    received_to: date | None = None,
    segment: str | None = None,
    status_filter: str | None = None,
    min_children: int = 2,
    limit: int = 1000,
) -> list[IntakeFamilySummaryRow]:
    if received_from is not None and received_to is not None and received_from > received_to:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'received_from' must be before 'received_to'",
        )
    stmt = (
        select(TypeformIntake, TypeformFormConfig)
        .outerjoin(TypeformFormConfig, TypeformFormConfig.id == TypeformIntake.form_config_id)
        .order_by(TypeformIntake.received_at.desc(), TypeformIntake.created_at.desc())
        .limit(limit)
    )
    if school_year_label:
        stmt = stmt.where(TypeformIntake.detected_school_year == school_year_label)
    if received_from is not None:
        start_local, _ = _day_bounds(received_from)
        stmt = stmt.where(TypeformIntake.received_at >= start_local)
    if received_to is not None:
        _, end_local = _day_bounds(received_to)
        stmt = stmt.where(TypeformIntake.received_at < end_local)
    if segment:
        stmt = stmt.where(TypeformIntake.detected_segment.ilike(segment.strip()))
    if status_filter:
        stmt = stmt.where(TypeformIntake.intake_status.ilike(status_filter.strip()))
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                TypeformIntake.source_form_id.ilike(like),
                TypeformIntake.source_response_id.ilike(like),
                TypeformIntake.detected_location.ilike(like),
                TypeformIntake.detected_segment.ilike(like),
                cast(TypeformIntake.normalized_payload_json, Text).ilike(like),
                cast(TypeformIntake.simplified_response_json, Text).ilike(like),
            )
        )

    grouped: dict[str, IntakeFamilySummaryRow] = {}
    latest_received: dict[str, datetime] = {}
    for intake, config in db.execute(stmt).all():
        normalized = _json_object(intake.normalized_payload_json)
        customer_type = _text(normalized.get("customer_type")).casefold()
        child_name = _display_name(
            normalized.get("child_first_name"),
            normalized.get("child_last_name"),
            fallback="",
        )
        if customer_type == "adult" and not child_name:
            continue
        if not child_name:
            child_name = _display_name(
                normalized.get("student_first_name"),
                normalized.get("student_last_name"),
                fallback="Enfant sans nom",
            )
        family_key = _intake_family_key(normalized)
        if not family_key:
            family_key = f"intake:{intake.id}"

        parent_name = _display_name(
            normalized.get("parent_first_name"),
            normalized.get("parent_last_name"),
            fallback="",
        ) or None
        parent_email = _text(normalized.get("parent_email")) or None
        parent_phone = _text(normalized.get("parent_phone")) or None
        family_label = parent_name or parent_email or parent_phone or "Famille sans contact"
        bucket = grouped.get(family_key)
        if bucket is None:
            bucket = IntakeFamilySummaryRow(
                family_key=family_key,
                family_label=family_label,
                parent_name=parent_name,
                parent_email=parent_email,
                parent_phone=parent_phone,
                intake_count=0,
                children=[],
            )
            grouped[family_key] = bucket
            latest_received[family_key] = intake.received_at
        else:
            bucket.parent_name = bucket.parent_name or parent_name
            bucket.parent_email = bucket.parent_email or parent_email
            bucket.parent_phone = bucket.parent_phone or parent_phone
            if bucket.family_label == "Famille sans contact" and family_label != "Famille sans contact":
                bucket.family_label = family_label
            if intake.received_at > latest_received[family_key]:
                latest_received[family_key] = intake.received_at

        bucket.intake_count += 1
        bucket.children.append(
            IntakeFamilyChildSummary(
                intake_id=intake.id,
                received_at=intake.received_at,
                source_form_id=intake.source_form_id,
                source_form_label=_form_label(config),
                child_name=child_name,
                segment=intake.detected_segment,
                status=intake.intake_status,
                course_1=_format_intake_course_1(normalized),
                course_2=_format_intake_course_2(normalized),
                solfege=_format_intake_solfege(normalized),
                masterclass=_format_intake_masterclass(normalized),
                pass_recup=_format_intake_pass_recup(normalized),
            )
        )

    families = [
        family
        for family in grouped.values()
        if len({child.child_name.casefold() for child in family.children}) >= min_children
    ]
    for family in families:
        family.children.sort(key=lambda child: (child.child_name.casefold(), child.received_at))
    families.sort(
        key=lambda family: (
            -len(family.children),
            family.family_label.casefold(),
            latest_received.get(family.family_key, datetime.min.replace(tzinfo=timezone.utc)),
        )
    )
    return families


@router.get("/intake-families", response_model=list[IntakeFamilySummaryRow])
def report_intake_families(
    q: str | None = None,
    school_year_label: str | None = None,
    received_from: date | None = None,
    received_to: date | None = None,
    segment: str | None = None,
    status_filter: str | None = Query(default=None, alias="status"),
    min_children: int = Query(default=2, ge=1, le=20),
    limit: int = Query(default=1000, ge=1, le=5000),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[IntakeFamilySummaryRow]:
    return _build_intake_family_summary_rows(
        db,
        q=q,
        school_year_label=school_year_label,
        received_from=received_from,
        received_to=received_to,
        segment=segment,
        status_filter=status_filter,
        min_children=min_children,
        limit=limit,
    )


@router.get("/check-deposits-due")
def export_check_deposits_due_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000, le=2100),
    legal_entity_id: UUID = Query(...),
    file_format: str = Query(default="pdf", pattern="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    legal_entity, rows = _check_deposit_report_rows(
        db,
        month=month,
        year=year,
        legal_entity_id=legal_entity_id,
    )
    entity_slug = _normalize_token(legal_entity.name) or "entite"
    filename_base = f"cheques-a-deposer-{year}-{month:02d}-{entity_slug}"
    if file_format == "xlsx":
        content = _render_check_deposit_report_xlsx(
            rows=rows,
            month=month,
            year=year,
            legal_entity=legal_entity,
        )
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )
    content = _render_check_deposit_report_pdf(
        rows=rows,
        month=month,
        year=year,
        legal_entity=legal_entity,
    )
    return Response(
        content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )


def _generated_report_html(row: GeneratedReport) -> str:
    content = _json_object(row.content_json)
    families = _json_list(content.get("families"))
    items = _json_list(content.get("items"))
    criteria = _json_object(row.criteria_json)
    title = html.escape(row.report_label)
    generated_at = row.created_at.astimezone(ADMIN_COMMUNICATION_TIMEZONE).strftime("%d/%m/%Y %H:%M")
    period_label = html.escape(_report_period_label(row.period_start, row.period_end))
    note = html.escape(row.note or "-")
    criteria_parts = [
        f"{html.escape(str(key))}: {html.escape(str(value))}"
        for key, value in criteria.items()
        if value not in (None, "", [])
    ]
    criteria_html = " | ".join(criteria_parts) or "-"
    blocks: list[str] = []
    if row.report_type == "expired-quotes":
        item_rows: list[str] = []
        for raw_item in items:
            item = _json_object(raw_item)
            item_rows.append(
                "<tr>"
                f"<td class='code'>{html.escape(_text(item.get('quote_number')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('student_name')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('contact_name')) or '-')}<br><span class='small'>{html.escape(_text(item.get('contact_email')) or '-')}</span></td>"
                f"<td>{html.escape(_text(item.get('effective_status_label')) or _text(item.get('status')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('effective_date'))[:10] or '-')}</td>"
                f"<td>{html.escape(_text(item.get('expires_at'))[:10] or '-')}</td>"
                f"<td>{html.escape(_text(item.get('total_ttc')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('location')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('planning')) or '-')}</td>"
                "</tr>"
            )
        table_html = (
            "<table><colgroup>"
            "<col style='width:14%'><col style='width:13%'><col style='width:16%'><col style='width:12%'>"
            "<col style='width:9%'><col style='width:9%'><col style='width:9%'><col style='width:9%'><col style='width:9%'>"
            "</colgroup><thead><tr>"
            "<th>Devis</th><th>Eleve</th><th>Contact</th><th>Statut sortie</th><th>Date statut</th>"
            "<th>Expiration prevue</th><th>Total</th><th>Lieu</th><th>Planning</th>"
            "</tr></thead><tbody>"
            f"{''.join(item_rows)}"
            "</tbody></table>"
        ) if item_rows else "<p>Aucun devis expire pour cette periode.</p>"
        blocks.append(table_html)
        return (
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>"
            "@page { size: A4 landscape; margin: 14mm; }"
            "body { font-family: Arial, sans-serif; color: #222; font-size: 9pt; }"
            "h1 { font-size: 20pt; margin: 0 0 8px; }"
            ".meta { color: #555; margin-bottom: 14px; }"
            "table { width: 100%; border-collapse: collapse; table-layout: fixed; margin-top: 8px; }"
            "th, td { border: 1px solid #ccd3dd; padding: 5px; vertical-align: top; }"
            "th { background: #eef2f6; text-align: left; }"
            "td { word-wrap: break-word; overflow-wrap: break-word; }"
            ".small { color: #596579; font-size: 8pt; font-weight: normal; }"
            ".code { font-size: 7.6pt; line-height: 1.25; }"
            "</style></head><body>"
            f"<h1>{title}</h1>"
            f"<p class='meta'>Genere le {generated_at} | Periode: {period_label} | Format: PDF | Note: {note}</p>"
            f"<p class='meta'>Criteres: {criteria_html}</p>"
            f"{''.join(blocks)}"
            "</body></html>"
        )

    if row.report_type == "overdue-invoices":
        item_rows = []
        for raw_item in items:
            item = _json_object(raw_item)
            contact_bits = [
                _text(item.get("client_email")),
                _text(item.get("client_phone")),
            ]
            contact = "<br>".join(html.escape(bit) for bit in contact_bits if bit) or "-"
            delay = _text(item.get("days_overdue"))
            delay_label = f"{delay} j" if delay and delay != "0" else "Echeance atteinte"
            item_rows.append(
                "<tr>"
                f"<td class='code'>{html.escape(_text(item.get('invoice_number')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('client_name')) or '-')}<br><span class='small'>{contact}</span></td>"
                f"<td>{html.escape(_text(item.get('billing_entity')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('issued_date')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('due_date')) or '-')}</td>"
                f"<td>{html.escape(delay_label)}</td>"
                f"<td>{html.escape(_text(item.get('total_due')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('amount_paid')) or '-')}</td>"
                f"<td>{html.escape(_text(item.get('last_contact_at')) or '-')}</td>"
                "</tr>"
            )
        table_html = (
            "<table><colgroup>"
            "<col style='width:13%'><col style='width:19%'><col style='width:12%'><col style='width:9%'>"
            "<col style='width:9%'><col style='width:10%'><col style='width:10%'><col style='width:9%'><col style='width:9%'>"
            "</colgroup><thead><tr>"
            "<th>Facture</th><th>Client</th><th>Entite</th><th>Emission</th><th>Echeance</th>"
            "<th>Retard</th><th>Montant du</th><th>Deja paye</th><th>Dernier contact</th>"
            "</tr></thead><tbody>"
            f"{''.join(item_rows)}"
            "</tbody></table>"
        ) if item_rows else "<p>Aucune facture echue non payee pour cette periode.</p>"
        blocks.append(table_html)
        return (
            "<!doctype html><html><head><meta charset='utf-8'>"
            "<style>"
            "@page { size: A4 landscape; margin: 14mm; }"
            "body { font-family: Arial, sans-serif; color: #222; font-size: 9pt; }"
            "h1 { font-size: 20pt; margin: 0 0 8px; }"
            ".meta { color: #555; margin-bottom: 14px; }"
            "table { width: 100%; border-collapse: collapse; table-layout: fixed; margin-top: 8px; }"
            "th, td { border: 1px solid #ccd3dd; padding: 5px; vertical-align: top; }"
            "th { background: #eef2f6; text-align: left; }"
            "td { word-wrap: break-word; overflow-wrap: break-word; }"
            ".small { color: #596579; font-size: 8pt; font-weight: normal; }"
            ".code { font-size: 7.6pt; line-height: 1.25; }"
            "</style></head><body>"
            f"<h1>{title}</h1>"
            f"<p class='meta'>Genere le {generated_at} | Echeances: {period_label} | Format: PDF | Note: {note}</p>"
            f"<p class='meta'>Criteres: {criteria_html}</p>"
            f"{''.join(blocks)}"
            "</body></html>"
        )

    report_rows = [
        ("course_1", "Cours 1"),
        ("course_2", "Cours 2"),
        ("solfege", "Solfege"),
        ("masterclass", "MasterClass"),
    ] if row.report_type == "quote-families" else [
        ("course_1", "Cours 1"),
        ("course_2", "Cours 2"),
        ("solfege", "Solfege"),
        ("masterclass", "Masterclass"),
        ("pass_recup", "Pass recup"),
    ]
    for family_raw in families:
        family = _json_object(family_raw)
        children = _json_list(family.get("children"))
        header_cells: list[str] = []
        for child in children:
            child_obj = _json_object(child)
            child_name = html.escape(_text(child_obj.get("child_name")) or "-")
            if row.report_type == "quote-families":
                quote_bits = [
                    _text(child_obj.get("quote_number")),
                    _text(child_obj.get("status")),
                    _text(child_obj.get("total_ttc")),
                ]
                meta = " · ".join(bit for bit in quote_bits if bit)
                header_cells.append(f"<th>{child_name}<br><span class='small'>{html.escape(meta)}</span></th>")
            else:
                header_cells.append(f"<th>{child_name}</th>")
        headers = "".join(header_cells)
        rows: list[str] = []
        for key, label in report_rows:
            cells = "".join(
                f"<td>{html.escape(_text(_json_object(child).get(key)) or '-')}</td>"
                for child in children
            )
            rows.append(f"<tr><th>{label}</th>{cells}</tr>")
        contact = _text(family.get("parent_email")) or _text(family.get("parent_phone")) or "-"
        count_label = f"{len(children)} devis" if row.report_type == "quote-families" else f"{len(children)} demande(s)"
        blocks.append(
            "<section class='family'>"
            f"<h2>{html.escape(_text(family.get('family_label')) or 'Famille')}</h2>"
            f"<p>{html.escape(count_label)} | {html.escape(contact)}</p>"
            "<table><thead><tr><th>Element</th>"
            f"{headers}</tr></thead><tbody>{''.join(rows)}</tbody></table>"
            "</section>"
        )
    if not blocks:
        blocks.append("<p>Aucune donnee pour ce rapport.</p>")
    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<style>"
        "@page { size: A4 landscape; margin: 18mm; }"
        "body { font-family: Arial, sans-serif; color: #222; font-size: 10pt; }"
        "h1 { font-size: 20pt; margin: 0 0 8px; } h2 { font-size: 13pt; margin: 18px 0 4px; }"
        ".meta { color: #555; margin-bottom: 14px; }"
        "table { width: 100%; border-collapse: collapse; margin-top: 8px; }"
        "th, td { border: 1px solid #ccd3dd; padding: 6px; vertical-align: top; }"
        "th { background: #eef2f6; text-align: left; } .family { page-break-inside: avoid; }"
        ".small { color: #596579; font-size: 8pt; font-weight: normal; }"
        "</style></head><body>"
        f"<h1>{title}</h1>"
        f"<p class='meta'>Genere le {generated_at} | Periode: {period_label} | Format: PDF | Note: {note}</p>"
        f"<p class='meta'>Criteres: {criteria_html}</p>"
        f"{''.join(blocks)}"
        "</body></html>"
    )


def _render_generated_report_pdf(row: GeneratedReport) -> bytes:
    if row.report_type == "check-deposits":
        return _render_stored_check_deposit_report_pdf(row)
    output = io.BytesIO()
    result = pisa.CreatePDF(src=_generated_report_html(row), dest=output, encoding="utf-8")
    if result.err:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Generation PDF impossible")
    return output.getvalue()


def _generated_report_filename(row: GeneratedReport, extension: str) -> str:
    slug = _normalize_token(row.report_label) or _normalize_token(row.report_type) or "rapport"
    created = row.created_at.astimezone(ADMIN_COMMUNICATION_TIMEZONE).strftime("%Y%m%d-%H%M")
    return f"{slug}-{created}.{extension}"


def _render_generated_report_download(row: GeneratedReport) -> tuple[bytes, str, str]:
    if row.file_format.upper() == "XLSX":
        if row.report_type == "check-deposits":
            return (
                _render_stored_check_deposit_report_xlsx(row),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                _generated_report_filename(row, "xlsx"),
            )
        if row.report_type == "material-forecast":
            return (
                _render_material_forecast_report_xlsx(row),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                _generated_report_filename(row, "xlsx"),
            )
        else:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Export Excel indisponible pour ce rapport")
    return _render_generated_report_pdf(row), "application/pdf", _generated_report_filename(row, "pdf")


@router.get("/generated", response_model=list[GeneratedReportOut])
def list_generated_reports(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[GeneratedReportOut]:
    rows = db.scalars(select(GeneratedReport).order_by(GeneratedReport.created_at.desc()).limit(500)).all()
    return [_generated_report_out(row) for row in rows]


@router.post("/generated", response_model=GeneratedReportOut, status_code=status.HTTP_201_CREATED)
def create_generated_report(
    payload: GeneratedReportCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
) -> GeneratedReportOut:
    report_type = payload.report_type.strip()
    report_label = REPORT_TYPE_LABELS.get(report_type, report_type)
    criteria = dict(payload.criteria or {})
    period_start = payload.period_start or _parse_report_date(criteria.get("received_from"))
    period_end = payload.period_end or _parse_report_date(criteria.get("received_to"))
    if period_start is not None and period_end is not None and period_start > period_end:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="'period_start' must be before 'period_end'")

    file_format = _text(criteria.get("file_format")).upper() or "PDF"
    if file_format not in {"PDF", "XLSX"}:
        file_format = "PDF"
    if report_type == "material-forecast":
        file_format = "XLSX"
    content: dict[str, object] = {"items": []}
    row_count = 0
    if report_type == "intake-families":
        try:
            min_children = int(criteria.get("min_children") or 2)
        except (TypeError, ValueError):
            min_children = 2
        min_children = max(1, min(20, min_children))
        families = _build_intake_family_summary_rows(
            db,
            q=_text(criteria.get("q")) or None,
            school_year_label=_text(criteria.get("school_year_label")) or None,
            received_from=period_start,
            received_to=period_end,
            segment=_text(criteria.get("segment")) or None,
            status_filter=_text(criteria.get("status")) or None,
            min_children=min_children,
            limit=5000,
        )
        content = {"families": [family.model_dump(mode="json") for family in families]}
        row_count = len(families)
    elif report_type == "quote-families":
        try:
            min_children = int(criteria.get("min_children") or 2)
        except (TypeError, ValueError):
            min_children = 2
        min_children = max(1, min(20, min_children))
        families = _build_quote_family_summary_rows(
            db,
            q=_text(criteria.get("q")) or None,
            school_year_label=_text(criteria.get("school_year_label")) or None,
            received_from=period_start,
            received_to=period_end,
            status_filter=_text(criteria.get("status")) or None,
            min_children=min_children,
            limit=5000,
        )
        content = {"families": families}
        row_count = len(families)
    elif report_type == "expired-quotes":
        rows = _build_expired_quote_rows(
            db,
            q=_text(criteria.get("q")) or None,
            school_year_label=_text(criteria.get("school_year_label")) or None,
            expired_from=period_start,
            expired_to=period_end,
            status_filter=_text(criteria.get("status")) or None,
            limit=5000,
        )
        content = {"items": rows}
        row_count = len(rows)
    elif report_type == "overdue-invoices":
        effective_period_end = period_end or datetime.now(ADMIN_COMMUNICATION_TIMEZONE).date()
        rows = _build_overdue_invoice_rows(
            db,
            q=_text(criteria.get("q")) or None,
            due_from=period_start,
            due_to=effective_period_end,
            limit=5000,
        )
        period_end = effective_period_end
        content = {"items": rows}
        row_count = len(rows)
    elif report_type == "check-deposits":
        try:
            month = int(criteria.get("month") or 0)
            year = int(criteria.get("year") or 0)
        except (TypeError, ValueError):
            month = 0
            year = 0
        if month < 1 or month > 12 or year < 2000 or year > 2100:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Mois de depot invalide")
        try:
            legal_entity_id = UUID(_text(criteria.get("legal_entity_id")))
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Entite legale invalide") from exc
        legal_entity, rows = _check_deposit_report_rows(
            db,
            month=month,
            year=year,
            legal_entity_id=legal_entity_id,
        )
        period_start = date(year, month, 1)
        period_end = date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1) - timedelta(days=1)
        report_label = f"{report_label} - {_month_year_label(month, year)} - {legal_entity.name}"
        content = {
            "items": [_check_deposit_report_row_json(item) for item in rows],
            "month": month,
            "year": year,
            "legal_entity_id": str(legal_entity.id),
            "legal_entity_name": legal_entity.name,
        }
        row_count = len(rows)
    elif report_type == "material-forecast":
        school_year_label = _text(criteria.get("school_year_label")) or None
        if school_year_label and period_start is None and period_end is None:
            period_start, period_end = _material_report_school_year_bounds(school_year_label)
        content = _build_material_forecast_report(
            db,
            school_year_label=school_year_label,
            approved_from=_parse_report_date(criteria.get("received_from")),
            approved_to=_parse_report_date(criteria.get("received_to")),
            status_filter=_text(criteria.get("status")) or "approved",
            q=_text(criteria.get("q")) or None,
        )
        row_count = len(_json_list(content.get("summary_rows")))
        if school_year_label:
            report_label = f"{report_label} - {school_year_label}"

    row = GeneratedReport(
        report_type=report_type,
        report_label=report_label,
        file_format=file_format,
        period_start=period_start,
        period_end=period_end,
        note=_text(payload.note) or None,
        criteria_json=criteria,
        content_json=content,
        row_count=row_count,
        created_by_user_id=current_user.id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _generated_report_out(row)


@router.get("/generated/{report_id}/pdf")
def download_generated_report_pdf(
    report_id: UUID,
    inline: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    row = db.get(GeneratedReport, report_id)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rapport introuvable")
    filename = f"rapport-{row.report_type}-{row.created_at.strftime('%Y%m%d')}.pdf".replace('"', "")
    disposition = "inline" if inline else "attachment"
    return Response(
        content=_render_generated_report_pdf(row),
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
    )


@router.get("/generated/{report_id}/download")
def download_generated_report(
    report_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    row = db.get(GeneratedReport, report_id)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rapport introuvable")
    content, media_type, filename = _render_generated_report_download(row)
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename.replace(chr(34), "")}"'},
    )


@router.delete("/generated", status_code=status.HTTP_204_NO_CONTENT)
def delete_generated_reports(
    report_ids: list[UUID] = Query(default=[]),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    if not report_ids:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Aucun rapport selectionne")
    rows = db.scalars(select(GeneratedReport).where(GeneratedReport.id.in_(report_ids))).all()
    found_ids = {row.id for row in rows}
    missing_ids = [str(report_id) for report_id in report_ids if report_id not in found_ids]
    if missing_ids:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rapport introuvable")
    for row in rows:
        db.delete(row)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.delete("/generated/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_generated_report(
    report_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    row = db.get(GeneratedReport, report_id)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rapport introuvable")
    db.delete(row)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/sap/{year}/csv")
def report_sap_csv(
    year: int = Path(..., ge=2000, le=2100),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> Response:
    services_entity_id = db.scalar(
        select(LegalEntity.id).where(func.upper(func.trim(LegalEntity.name)) == "PIANO ACADEMIE SERVICES").limit(1)
    )
    if services_entity_id is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Legal entity 'PIANO ACADEMIE SERVICES' not found",
        )

    period_start = datetime(year, 1, 1, tzinfo=timezone.utc)
    period_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    rows = db.execute(
        select(ClientInvoiceLine, ClientNoteEntry, User)
        .join(ClientNoteEntry, ClientNoteEntry.id == ClientInvoiceLine.note_id)
        .join(User, User.id == ClientInvoiceLine.user_id)
        .where(
            ClientInvoiceLine.seller_legal_entity_id == services_entity_id,
            ClientInvoiceLine.occurred_at >= period_start,
            ClientInvoiceLine.occurred_at < period_end,
        )
        .order_by(
            func.lower(func.coalesce(User.last_name, "")),
            func.lower(func.coalesce(User.first_name, "")),
            func.lower(User.email),
            ClientInvoiceLine.occurred_at.asc(),
            ClientInvoiceLine.id.asc(),
        )
    ).all()

    grouped: dict[str, dict[str, object]] = {}
    for line, note, client in rows:
        client_id = str(client.id)
        bucket = grouped.setdefault(
            client_id,
            {
                "client_id": client_id,
                "client_name": _client_name(client),
                "client_email": client.email,
                "service_address": _service_address(client),
                "total_paid_ttc": Decimal("0.00"),
                "details": [],
            },
        )
        invoice_number, payment_status = _invoice_fields_from_note_message(note.message)
        line_total = Decimal(line.total_incl_vat or 0).quantize(Decimal("0.01"))
        bucket["total_paid_ttc"] = (Decimal(bucket["total_paid_ttc"]) + line_total).quantize(Decimal("0.01"))
        details_bucket = bucket["details"]
        if not isinstance(details_bucket, list):
            details_bucket = []
            bucket["details"] = details_bucket
        details_bucket.append(
            {
                "line_occurred_at": line.occurred_at,
                "invoice_number": invoice_number or "",
                "label": line.label or "",
                "total_incl_vat": line_total,
                "currency": (line.currency or "").upper(),
                "note_id": str(line.note_id),
                "payment_status": payment_status or "",
            }
        )

    output_rows: list[dict[str, str]] = []
    sorted_clients = sorted(
        grouped.values(),
        key=lambda row: (
            str(row["client_name"]).casefold(),
            str(row["client_email"]).casefold(),
        ),
    )
    for bucket in sorted_clients:
        details_raw = bucket.get("details")
        details = details_raw if isinstance(details_raw, list) else []
        details.sort(
            key=lambda row: (
                row.get("line_occurred_at")
                if isinstance(row, dict) and isinstance(row.get("line_occurred_at"), datetime)
                else period_start
            )
        )
        statuses = sorted({str(row["payment_status"]).strip().upper() for row in details if str(row["payment_status"]).strip()})
        summary_status = statuses[0] if len(statuses) == 1 else ("MIXED" if statuses else "")
        total_paid = Decimal(bucket["total_paid_ttc"]).quantize(Decimal("0.01"))

        output_rows.append(
            {
                "row_type": "SUMMARY",
                "year": str(year),
                "client_id": str(bucket["client_id"]),
                "client_name": str(bucket["client_name"]),
                "client_email": str(bucket["client_email"]),
                "total_paid_ttc": f"{total_paid:.2f}",
                "line_occurred_at": "",
                "invoice_number": "",
                "label": "TOTAL_CLIENT",
                "total_incl_vat": "",
                "currency": "",
                "service_address": str(bucket["service_address"]),
                "note_id": "",
                "payment_status": summary_status,
            }
        )
        for detail in details:
            detail_occurred_at = detail.get("line_occurred_at")
            detail_total = detail.get("total_incl_vat")
            try:
                detail_total_decimal = Decimal(str(detail_total)).quantize(Decimal("0.01"))
            except Exception:
                detail_total_decimal = Decimal("0.00")
            output_rows.append(
                {
                    "row_type": "DETAIL",
                    "year": str(year),
                    "client_id": str(bucket["client_id"]),
                    "client_name": str(bucket["client_name"]),
                    "client_email": str(bucket["client_email"]),
                    "total_paid_ttc": f"{total_paid:.2f}",
                    "line_occurred_at": (detail_occurred_at.isoformat() if isinstance(detail_occurred_at, datetime) else ""),
                    "invoice_number": str(detail["invoice_number"]),
                    "label": str(detail["label"]),
                    "total_incl_vat": f"{detail_total_decimal:.2f}",
                    "currency": str(detail["currency"]),
                    "service_address": str(bucket["service_address"]),
                    "note_id": str(detail["note_id"]),
                    "payment_status": str(detail["payment_status"]),
                }
            )

    csv_columns = [
        "row_type",
        "year",
        "client_id",
        "client_name",
        "client_email",
        "total_paid_ttc",
        "line_occurred_at",
        "invoice_number",
        "label",
        "total_incl_vat",
        "currency",
        "service_address",
        "note_id",
        "payment_status",
    ]
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=csv_columns)
    writer.writeheader()
    writer.writerows(output_rows)

    file_name = f"sap_services_export_{year}.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/communications", response_model=CommunicationReportPageOut)
def report_communications(
    channel: CommunicationChannel | None = Query(default=None),
    period: CommunicationPeriod = Query(default=CommunicationPeriod.TODAY),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=25, le=100),
    q: str | None = Query(default=None),
    communication_type: str | None = Query(default=None),
    occurred_on: date | None = Query(default=None),
    professor_id: UUID | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> CommunicationReportPageOut:
    if per_page not in (25, 50, 100):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="per_page must be one of: 25, 50, 100",
        )

    now_utc = datetime.now(timezone.utc)
    _archive_communications_older_than_one_year(db, now_utc)

    normalized_type = (communication_type or "").strip().upper()
    search = (q or "").strip()
    period_bounds = _communication_period_bounds(period, now_utc)
    if occurred_on is not None:
        period_bounds = _day_bounds(occurred_on)

    filters: list = []
    if channel is not None:
        filters.append(CommunicationLog.channel == channel.value)
    if period_bounds is not None:
        start_at, end_at = period_bounds
        filters.append(CommunicationLog.occurred_at >= start_at)
        filters.append(CommunicationLog.occurred_at < end_at)

    if normalized_type and normalized_type != "ALL":
        filters.append(func.upper(CommunicationLog.communication_type) == normalized_type)
    if professor_id is not None:
        filters.append(CommunicationLog.professor_id == professor_id)
    if search:
        pattern = f"%{search.lower()}%"
        filters.append(
            or_(
                func.lower(CommunicationLog.subject).like(pattern),
                func.lower(CommunicationLog.sender_label).like(pattern),
                func.lower(CommunicationLog.recipient).like(pattern),
                func.lower(CommunicationLog.content).like(pattern),
                func.lower(CommunicationLog.source).like(pattern),
            )
        )

    count_stmt = select(func.count()).select_from(CommunicationLog)
    if filters:
        count_stmt = count_stmt.where(*filters)
    total = int(db.scalar(count_stmt) or 0)
    total_pages = max(1, (total + per_page - 1) // per_page)
    current_page = min(page, total_pages)
    offset = (current_page - 1) * per_page

    data_stmt = select(CommunicationLog)
    if filters:
        data_stmt = data_stmt.where(*filters)
    data_stmt = data_stmt.order_by(CommunicationLog.occurred_at.desc()).offset(offset).limit(per_page)
    rows = db.scalars(data_stmt).all()

    return CommunicationReportPageOut(
        items=[_communication_row_out(row) for row in rows],
        page=current_page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@router.get("/communications/filters", response_model=CommunicationFiltersOut)
def report_communication_filters(
    channel: CommunicationChannel | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> CommunicationFiltersOut:
    type_codes: set[str] = set()
    for code in KNOWN_COMMUNICATION_TYPES:
        type_codes.add(code)
    db_type_stmt = select(CommunicationLog.communication_type).distinct()
    if channel is not None:
        db_type_stmt = db_type_stmt.where(CommunicationLog.channel == channel.value)
    db_type_rows = db.scalars(db_type_stmt).all()
    for code in db_type_rows:
        normalized = (code or "").strip().upper()
        if normalized:
            type_codes.add(normalized)
    communication_types = [
        CommunicationTypeFilterOut(code=code, label=communication_type_label(code))
        for code in sorted(type_codes, key=lambda item: COMMUNICATION_TYPE_LABELS.get(item, item))
    ]

    professors = db.scalars(
        select(Professor)
        .where(Professor.active.is_(True))
        .order_by(Professor.first_name.asc(), Professor.last_name.asc())
    ).all()
    professor_options = [
        CommunicationProfessorFilterOut(
            id=professor.id,
            label=f"{(professor.first_name or '').strip()} {(professor.last_name or '').strip()}".strip() or professor.email,
        )
        for professor in professors
    ]

    return CommunicationFiltersOut(
        communication_types=communication_types,
        professors=professor_options,
    )
