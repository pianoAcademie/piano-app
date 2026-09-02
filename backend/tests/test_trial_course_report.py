from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import sys
import unittest
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.api.routes.reports import (
    _trial_attendance_label,
    _trial_conversion_status,
    _trial_courses_xlsx,
    _trial_detection_source,
    _trial_enrollment_evidence,
)
from app.models.catalog import BookingStatus
from app.schemas.report import TrialCourseReportRow


class TrialCourseReportTests(unittest.TestCase):
    def test_persisted_trial_marker_survives_subscription_and_status_changes(self) -> None:
        start_at = datetime(2026, 9, 10, 14, 0, tzinfo=timezone.utc)
        session = SimpleNamespace(title="Piano", start_at_utc=start_at)
        course_type = SimpleNamespace(name="Piano individuel", code="PIANO_30")
        booking = SimpleNamespace(
            is_trial_course=True,
            student_start_at_utc=None,
            client_plan_subscription_id=uuid4(),
        )
        user = SimpleNamespace(client_status="ACTIVE", first_course_at=start_at)

        self.assertEqual(_trial_detection_source(session, course_type, booking, user), "RESERVATION_ESSAI")

    def test_first_course_without_subscription_is_detected_after_conversion(self) -> None:
        start_at = datetime(2026, 9, 10, 14, 0, tzinfo=timezone.utc)
        session = SimpleNamespace(title="Piano", start_at_utc=start_at)
        course_type = SimpleNamespace(name="Piano individuel", code="PIANO_30")
        booking = SimpleNamespace(student_start_at_utc=None, client_plan_subscription_id=None)
        user = SimpleNamespace(client_status="ACTIVE", first_course_at=start_at)

        self.assertEqual(
            _trial_detection_source(session, course_type, booking, user),
            "PREMIER_COURS_HORS_ABONNEMENT",
        )

    def test_regular_subscribed_first_course_is_not_detected_as_trial(self) -> None:
        start_at = datetime(2026, 9, 10, 14, 0, tzinfo=timezone.utc)
        session = SimpleNamespace(title="Piano", start_at_utc=start_at)
        course_type = SimpleNamespace(name="Piano individuel", code="PIANO_30")
        booking = SimpleNamespace(student_start_at_utc=None, client_plan_subscription_id=uuid4())
        user = SimpleNamespace(client_status="ACTIVE", first_course_at=start_at)

        self.assertIsNone(_trial_detection_source(session, course_type, booking, user))

    def test_past_booked_trial_is_reported_as_attendance_missing(self) -> None:
        now = datetime(2026, 9, 11, 10, 0, tzinfo=timezone.utc)
        label = _trial_attendance_label(
            BookingStatus.BOOKED,
            session_start_at=now - timedelta(days=1),
            now=now,
        )
        self.assertEqual(label, "Presence non renseignee")

    def test_future_trial_with_account_only_is_not_marked_registered(self) -> None:
        now = datetime(2026, 9, 2, 10, 0, tzinfo=timezone.utc)
        self.assertEqual(
            _trial_conversion_status(
                session_start_at=now + timedelta(hours=2),
                session_end_at=now + timedelta(hours=2, minutes=30),
                now=now,
                attendance_status="BOOKED",
                client_kind="ADULT",
                enrollment_evidence=None,
                quote_status=None,
                has_intake=False,
            ),
            "Essai à venir - compte créé",
        )

    def test_past_trial_without_progress_is_marked_for_followup(self) -> None:
        now = datetime(2026, 9, 2, 18, 0, tzinfo=timezone.utc)
        self.assertEqual(
            _trial_conversion_status(
                session_start_at=now - timedelta(hours=2),
                session_end_at=now - timedelta(hours=1, minutes=30),
                now=now,
                attendance_status="ATTENDED",
                client_kind="ADULT",
                enrollment_evidence=None,
                quote_status=None,
                has_intake=False,
            ),
            "À relancer",
        )

    def test_child_is_registered_only_with_approved_quote(self) -> None:
        self.assertEqual(
            _trial_enrollment_evidence(
                client_kind="CHILD",
                quote_status="approved",
                subscriptions=[],
            ),
            (True, "Devis validé", "APPROVED_QUOTE"),
        )

    def test_adult_pending_subscription_is_not_registered(self) -> None:
        self.assertEqual(
            _trial_enrollment_evidence(
                client_kind="ADULT",
                quote_status=None,
                subscriptions=[("SUBSCRIPTION", "PENDING")],
            ),
            (False, "Non confirmée", None),
        )

    def test_adult_active_pack_is_registered_with_explicit_evidence(self) -> None:
        self.assertEqual(
            _trial_enrollment_evidence(
                client_kind="ADULT",
                quote_status=None,
                subscriptions=[("PACK", "ACTIVE")],
            ),
            (True, "Carnet actif", "ACTIVE_PACK"),
        )

    def test_xlsx_contains_detail_and_summary_sheets(self) -> None:
        start_at = datetime(2026, 9, 10, 14, 0, tzinfo=timezone.utc)
        row = TrialCourseReportRow(
            booking_id=uuid4(),
            session_id=uuid4(),
            session_start_at=start_at,
            session_end_at=start_at + timedelta(minutes=30),
            session_timezone="Europe/Paris",
            course_type_name="Initiation piano",
            course_format="PARTICULIER",
            location_id=uuid4(),
            location_name="Richelieu",
            professor_id=uuid4(),
            professor_name="Camille Prof",
            student_id=uuid4(),
            student_first_name="Lina",
            student_last_name="Martin",
            student_email="lina@example.test",
            parent_email="parent@example.test",
            attendance_status="ATTENDED",
            attendance_label="Present",
            internal_note="Tres bon contact",
            conversion_status="Inscrit - devis validé",
            account_status_label="Compte créé",
            client_kind="CHILD",
            client_status="ACTIVE",
            has_intake=True,
            intake_status_label="Reçu",
            intake_status="PROCESSED",
            intake_received_at=start_at - timedelta(days=2),
            quote_status="approved",
            quote_status_label="Validé",
            is_registered=True,
            enrollment_status_label="Devis validé",
            enrollment_evidence="APPROVED_QUOTE",
            trial_detection_source="STATUT_ESSAI",
        )

        workbook = load_workbook(BytesIO(_trial_courses_xlsx([row])))

        self.assertEqual(workbook.sheetnames, ["Cours d'essai", "Synthese"])
        detail = workbook["Cours d'essai"]
        self.assertEqual(detail["H2"].value, "Lina")
        self.assertEqual(detail["M2"].value, "Tres bon contact")
        self.assertEqual(detail["N2"].value, "Inscrit - devis validé")
        self.assertEqual(workbook["Synthese"]["B5"].value, 1)


if __name__ == "__main__":
    unittest.main()
