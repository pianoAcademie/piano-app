from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import sys
import unittest

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.api.routes.reports import _render_material_forecast_report_xlsx


class MaterialForecastI18nTests(unittest.TestCase):
    def _render(self, language: str):
        row = SimpleNamespace(
            content_json={
                "summary_rows": [
                    {
                        "site": "PARIS",
                        "site_label": "Paris",
                        "kind": "Partition",
                        "category_name": "Partitions",
                        "product_title": "Example",
                        "expected_total": 2,
                        "expected_direct": 1,
                        "expected_from_kits": 1,
                        "stock_quantity": 0,
                        "product_id": "product-1",
                    }
                ],
                "details": [
                    {
                        "site": "PARIS",
                        "site_label": "Paris",
                        "quote_number": "D-001",
                        "approved_at": "2026-07-17T08:00:00+00:00",
                        "student_name": "Student",
                        "source": "Kit inscription",
                        "kit_title": "Kit",
                        "quote_line_title": "Line",
                        "kind": "Partition",
                        "category_name": "Partitions",
                        "product_title": "Example",
                        "quantity": 1,
                        "product_id": "product-1",
                    }
                ],
            },
            criteria_json={"ui_language": language},
            created_at=datetime(2026, 7, 17, 8, 0, tzinfo=timezone.utc),
        )
        return load_workbook(BytesIO(_render_material_forecast_report_xlsx(row)))

    def test_renders_english_workbook_labels(self) -> None:
        workbook = self._render("en")

        self.assertIn("Paris details", workbook.sheetnames)
        self.assertEqual(workbook["Paris"]["A1"].value, "Sheet music, workbooks, and note-card supplies - Paris")
        self.assertEqual(workbook["Paris"]["B5"].value, "Type")
        self.assertEqual(workbook["Paris"]["B6"].value, "Sheet music")
        self.assertEqual(workbook["Paris details"]["E4"].value, "Enrolment kit")

    def test_keeps_french_workbook_labels(self) -> None:
        workbook = self._render("fr")

        self.assertIn("Detail Paris", workbook.sheetnames)
        self.assertEqual(workbook["Paris"]["A1"].value, "Approvisionnement partitions, cahiers et jeux de notes - Paris")
        self.assertEqual(workbook["Paris"]["B5"].value, "Nature")


if __name__ == "__main__":
    unittest.main()
